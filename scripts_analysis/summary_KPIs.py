import os
import time
from datetime import datetime
from pathlib import Path

import openpyxl
import pandas as pd
import pypsa
from openpyxl.utils.dataframe import dataframe_to_rows

from common import log


def load_framework(framework_path):
    """
    Load an Excel framework from a specified path and convert the data
    type of the 'Subcategory' column to string.

    Parameters
    ----------
    framework_path : str
        The file path to the Excel framework file to be loaded.

    Returns
    -------
    pandas.DataFrame
        A pandas DataFrame containing the loaded Excel file with the
        'Subcategory' column converted to string.
    """
    # Load the Excel framework into which the KPIs are inserted later
    framework_df = pd.read_excel(framework_path)

    # Convert the data type of the 'Subcategory' column to string
    # for better data handling
    framework_df["Subcategory"] = framework_df["Subcategory"].astype(str)

    return framework_df


def get_bus_slice(
    EU27_countries, identifier, interesting_countries, interesting_nodes
):
    """
    Calculate the length of the bus slice needed to find identifiers
    and put identifiers into list.

    Parameters
    ----------
    EU27_countries : list of str
        List of EU27 country codes.
    identifier : str
        String of country, cluster, 'All_countries' or 'EU27'.
    interesting_countries : list of str
        List of country codes.
    interesting_nodes : list of str
        List of node identifiers.

    Returns
    -------
    slice
        Length of the relevant string identifying the region.
    list of str
        List of all interesting regions.
    """
    if identifier in interesting_countries:
        if identifier == "EU27":
            identifier_list = EU27_countries
        elif identifier == "All_Countries":
            identifier_list = [
                k for k in interesting_countries if k != "All_Countries"
            ]  # Avoid recursion
        else:
            identifier_list = [identifier]
        bus_slice = slice(2)  # Slice first two characters for country codes
    else:
        if identifier == "All_Nodes":
            identifier_list = [
                k for k in interesting_nodes if k != "All_Nodes"
            ]  # Avoid recursion
        else:
            identifier_list = [identifier]
        bus_slice = slice(5)  # Slice first five characters for nodes
    return bus_slice, identifier_list


def calculate_flh_links(
    snapshot_weightings,
    interesting_countries,
    EU27_countries,
    interesting_nodes,
    network,
    filter_carrier=["H2 Electrolysis"],
):
    """
    This function combines input lists of countries and nodes for
    processing, and calculates full load hours for any links given with
    the carriers given in the filter_carrier. Those carriers are not
    viewed separately but combined (eg. if you want OCGT, CCGT and
    allam cycle for gas power plants.

    Parameters
    ----------
    snapshot_weightings : pd.DataFrame
        DataFrame containing weightings for different snapshots in the
        analysis.
    interesting_countries : list of str
        List of country codes.
    EU27_countries : list of str
        List of EU27 country codes.
    interesting_nodes : list of str
        List of node identifiers.
    network : pypsa.Network
        The PyPSA Network object.
    filter_carrier : list, optional
        List of all the carriers that should be combined to calculate
        the full load hours., by default ["H2 Electrolysis"]

    Returns
    -------
    pd.Series
        A pandas Series containing the full load hours organized by
        identifier.
    """
    # Copy network and initialize variables.
    network_flh = network.copy()
    flhs = {}

    # Combine countries and nodes in one list for processing
    combined_identifiers = interesting_countries + interesting_nodes

    # Calculate energy metrics for each interesting identifier
    for identifier in combined_identifiers:
        bus_slice, identifier_list = get_bus_slice(
            EU27_countries,
            identifier,
            interesting_countries,
            interesting_nodes,
        )
        # Get index of components that have the right carrier and location.
        idx_flh = network_flh.links[
            network_flh.links.index.str[bus_slice].isin(identifier_list)
            & network_flh.links.carrier.isin(filter_carrier)
        ].index
        # Calculate the installed capacity.
        p_nom_opt = network_flh.links.loc[idx_flh, "p_nom_opt"].sum()
        # Calculate the energy demand/supply of the relevant components.
        dispatch = network_flh.links_t.p0.loc[:, idx_flh]
        energy = dispatch.mul(snapshot_weightings, axis=0).sum().sum()
        # Calculate the full load hours.
        flhs[identifier] = energy / p_nom_opt
    return pd.Series(flhs)


def load_and_process_RE_electricity(
    snapshot_weightings,
    interesting_countries,
    EU27_countries,
    interesting_nodes,
    network,
):
    """
    Processes renewable energy data by standardizing carrier names and
    calculating energy metrics for specified countries and nodes.

    This function replaces old carrier names with standardized ones,
    combines input lists of countries and nodes for processing, and
    calculates various energy metrics such as total generation,
    potential energy, capacity, full load hours, and curtailment.

    Parameters
    ----------
    snapshot_weightings : pd.DataFrame
        DataFrame containing weightings for different snapshots in the
        analysis.
    interesting_countries : list of str
        List of country codes.
    EU27_countries : list of str
        List of EU27 country codes.
    interesting_nodes : list of str
        List of node identifiers.
    network : pypsa.Network
        The PyPSA Network object.

    Returns
    -------
    pd.DataFrame
        A DataFrame containing the calculated energy metrics organized
        by carrier type and identifier.
    """
    n_re_elec = network.copy()
    # Map old carrier names to new ones for consistency
    carrier_map = {
        "solar": "Solar",
        "onwind": "Onwind",
        "solar rooftop": "Solar",
        "offwind-ac": "Offwind",
        "offwind-dc": "Offwind",
    }
    n_re_elec.generators.carrier.replace(carrier_map, inplace=True)

    # Define the carriers and initialize a dictionary for results
    interesting_carriers = ["Solar", "Onwind", "Offwind"]
    results = {}

    # Combine countries and nodes in one list for processing
    combined_identifiers = interesting_countries + interesting_nodes

    # Process data for each carrier type
    for carrier in interesting_carriers:
        carrier_generators = n_re_elec.generators[
            n_re_elec.generators.carrier == carrier
        ]
        total_generation_sum = {}
        max_energy_sum = {}
        opt_capacity_sum = {}
        full_load_hours_sum = {}
        curtailment_sum = {}

        # Calculate energy metrics for each interesting identifier
        for identifier in combined_identifiers:
            bus_slice, identifier_list = get_bus_slice(
                EU27_countries,
                identifier,
                interesting_countries,
                interesting_nodes,
            )

            idx_gen = carrier_generators[
                carrier_generators.index.str[bus_slice].isin(identifier_list)
            ].index

            # Compute total and maximum potential energy, optimized capacity,
            # full load hours and curtailment for RE
            total_generation = (
                n_re_elec.generators_t.p[idx_gen].mul(
                    snapshot_weightings, axis=0
                )
            ).sum().fillna(
                0
            ).sum() * 1e-6  # Convert to TWh
            max_generated_energy = (
                (
                    n_re_elec.generators_t.p_max_pu[idx_gen].mul(
                        snapshot_weightings, axis=0
                    )
                )
                * (n_re_elec.generators.loc[idx_gen, "p_nom_opt"])
            ).sum().fillna(0).sum() * 1e-6
            opt_capacity = (
                n_re_elec.generators.loc[idx_gen, "p_nom_opt"].sum() * 1e-3
            )  # Convert to GWh
            # Apply a filter to exclude values below 0.1. This threshold
            # aligns with the model's precision limit of one decimal place
            # and prevents the generation of unrealistic full load hours for
            # certain countries and years.
            full_load_hours = (
                (total_generation / opt_capacity) * 1000
                if (total_generation >= 0.01 and opt_capacity >= 0.1)
                else 0
            )
            curtailment = max_generated_energy - total_generation

            # Store results for the current identifier and carrier
            total_generation_sum[identifier] = total_generation
            max_energy_sum[identifier] = max_generated_energy
            opt_capacity_sum[identifier] = opt_capacity
            full_load_hours_sum[identifier] = full_load_hours
            curtailment_sum[identifier] = curtailment

        results[("total_generation", carrier)] = total_generation_sum
        results[("max_generated_energy", carrier)] = max_energy_sum
        results[("opt_capacity", carrier)] = opt_capacity_sum
        results[("full_load_hours", carrier)] = full_load_hours_sum
        results[("curtailment", carrier)] = curtailment_sum

    # Convert results to DataFrame, fill missing values with zero, and return
    result = pd.DataFrame(results)
    result_filled = result.fillna(0)
    return result_filled


def sum_generated_electricity(
    snapshot_weightings,
    interesting_countries,
    EU27_countries,
    interesting_nodes,
    network,
):
    """
    Calculates the total generated electricity across specified
    countries and nodes by summing generation from various sources.

    Parameters
    ----------
    snapshot_weightings : pd.DataFrame
        DataFrame containing weightings for different snapshots in the
        analysis.
    interesting_countries : list of str
        List of country codes.
    EU27_countries : list of str
        List of EU27 country codes.
    interesting_nodes : list of str
        List of node identifiers.
    network : pypsa.Network
        The PyPSA Network object.

    Returns
    -------
    pd.Series
        A pandas Series containing the sum of generated electricity,
        indexed by country or node identifiers.
    """
    n_gen_elec = network.copy()
    electricity_links = [
        "OCGT",
        "H2 Fuel Cell",
        "H2 turbine",
        "coal",
        "residential rural micro gas CHP",
        "services rural micro gas CHP",
        "residential urban decentral micro gas CHP",
        "services urban decentral micro gas CHP",
        "urban central gas CHP",
        "urban central gas CHP CC",
        "urban central solid biomass CHP",
        "urban central solid biomass CHP CC",
        "allam",
        "CCGT",
        "lignite",
        "nuclear",
        "oil",
    ]

    generation_sums = {}

    # Combine countries and nodes in one list for processing
    combined_identifiers = interesting_countries + interesting_nodes

    # Calculate generation sums for each interesting identifier
    for identifier in combined_identifiers:
        bus_slice, identifier_list = get_bus_slice(
            EU27_countries,
            identifier,
            interesting_countries,
            interesting_nodes,
        )

        # Find generators and links related to electricity production
        electricity_buses = n_gen_elec.buses[
            n_gen_elec.buses.carrier.isin(["AC", "low voltage"])
        ].index
        electricity_generators = n_gen_elec.generators[
            n_gen_elec.generators.bus.isin(electricity_buses)
        ].carrier.unique()

        idx_gen = n_gen_elec.generators[
            (n_gen_elec.generators.carrier.isin(electricity_generators))
            & (
                n_gen_elec.generators.index.str[bus_slice].isin(
                    identifier_list
                )
            )
        ].index
        generators_sum = (
            (
                n_gen_elec.generators_t.p[idx_gen].mul(
                    snapshot_weightings, axis=0
                )
            )
            .sum()
            .fillna(0)
            .sum()
        )

        idx_links = n_gen_elec.links[
            (n_gen_elec.links.carrier.isin(electricity_links))
            & (n_gen_elec.links.index.str[bus_slice].isin(identifier_list))
        ].index
        links_sum = (
            (
                -n_gen_elec.links_t.p1[idx_links].mul(
                    snapshot_weightings, axis=0
                )
            )
            .sum()
            .fillna(0)
            .sum()
        )

        # Account for electricity stored or released from storage units
        idx_storage = n_gen_elec.storage_units[
            (n_gen_elec.storage_units.carrier == "hydro")
            & (
                n_gen_elec.storage_units.index.str[bus_slice].isin(
                    identifier_list
                )
            )
        ].index
        storage_units_sum = (
            (
                n_gen_elec.storage_units_t.p[idx_storage].mul(
                    snapshot_weightings, axis=0
                )
            )
            .sum()
            .fillna(0)
            .sum()
        )

        # Combine all electricity generation and consumption figures
        # and remove small values to avoid wrong calculations
        # for the import share of electricity
        total_sum = (generators_sum + links_sum + storage_units_sum) * 1e-6
        generation_sums[identifier] = total_sum if total_sum >= 0.001 else 0

    # Convert results to a Pandas Series with named indices
    result_series = pd.Series(generation_sums)
    result_series.name = "generated_electricity"
    result_series.index.name = "country_or_node"

    return result_series


def calculate_net_electricity_import(
    snapshot_weightings,
    interesting_countries,
    EU27_countries,
    interesting_nodes,
    network,
):
    """
    Calculates the net electricity imports for specified countries or
    regions by analyzing inbound and outbound flows on both AC and DC
    networks.

    Parameters
    ----------
    snapshot_weightings : pd.DataFrame
        DataFrame containing weightings for different snapshots in the
        analysis.
    interesting_countries : list of str
        List of country codes.
    EU27_countries : list of str
        List of EU27 country codes.
    interesting_nodes : list of str
        List of node identifiers.
    network : pypsa.Network
        The PyPSA Network object.

    Returns
    -------
    pd.Series
        A pandas Series containing the net imports of electricity for
        each country or node. Values near zero are set to zero to avoid
        misrepresentation of import shares due to rounding errors.
    """
    n_net_elec = network.copy()
    import_sums = {}

    # Combine countries and nodes in one list for processing
    combined_identifiers = interesting_countries + interesting_nodes

    # Calculate net electricity imports for specified countries or regions
    for identifier in combined_identifiers:
        bus_slice, identifier_list = get_bus_slice(
            EU27_countries,
            identifier,
            interesting_countries,
            interesting_nodes,
        )

        # Identify outbound and inbound DC links
        idx_DC_out = n_net_elec.links[
            (n_net_elec.links.carrier == "DC")
            & (n_net_elec.links.bus0.str[bus_slice].isin(identifier_list))
            & ~(n_net_elec.links.bus1.str[bus_slice].isin(identifier_list))
        ].index
        idx_DC_in = n_net_elec.links[
            (n_net_elec.links.carrier == "DC")
            & (n_net_elec.links.bus1.str[bus_slice].isin(identifier_list))
            & ~(n_net_elec.links.bus0.str[bus_slice].isin(identifier_list))
        ].index
        e_DC_out = (
            n_net_elec.links_t.p0[idx_DC_out].mul(snapshot_weightings, axis=0)
        ).sum().sum() * 1e-6
        e_DC_in = (
            -n_net_elec.links_t.p1[idx_DC_in].mul(snapshot_weightings, axis=0)
        ).sum().sum() * 1e-6

        # Identify outbound and inbound AC lines
        idx_AC_out = n_net_elec.lines[
            (n_net_elec.lines.bus0.str[bus_slice].isin(identifier_list))
            & ~(n_net_elec.lines.bus1.str[bus_slice].isin(identifier_list))
        ].index
        idx_AC_in = n_net_elec.lines[
            (n_net_elec.lines.bus1.str[bus_slice].isin(identifier_list))
            & ~(n_net_elec.lines.bus0.str[bus_slice].isin(identifier_list))
        ].index
        e_AC_out = (
            n_net_elec.lines_t.p0[idx_AC_out].mul(snapshot_weightings, axis=0)
        ).sum().sum() * 1e-6
        e_AC_in = (
            -n_net_elec.lines_t.p1[idx_AC_in].mul(snapshot_weightings, axis=0)
        ).sum().sum() * 1e-6

        # Compute the net import sum by combining DC and AC flows
        import_sum = e_DC_in - e_DC_out + e_AC_in - e_AC_out

        # Store the net import result for each identifier and remove small
        # values to avoid wrong calculations for the
        # import share of electricity
        import_sums[identifier] = (
            import_sum if import_sum >= 0.001 or import_sum <= -0.001 else 0
        )

    # Convert results to a Pandas Series with named indices
    import_series = pd.Series(import_sums, name="net_import_electricity")
    import_series.index.name = "country_or_node"

    return import_series


def calculate_gross_electricity_consumption(
    snapshot_weightings,
    interesting_countries,
    EU27_countries,
    interesting_nodes,
    network,
):
    """
    Calculates the gross electricity consumption for specified
    countries or regions. This is achieved by summing the net
    electricity imports and generated electricity to determine the
    total consumption.

    Parameters
    ----------
    snapshot_weightings : pd.DataFrame
        DataFrame containing weightings for different snapshots in the
        analysis.
    interesting_countries : list of str
        List of country codes.
    EU27_countries : list of str
        List of EU27 country codes.
    interesting_nodes : list of str
        List of node identifiers.
    network : pypsa.Network
        The PyPSA Network object.

    Returns
    -------
    pd.Series
        A pandas Series containing the gross electricity consumption
        for each country or node.

    """
    n_gross_elec = network.copy()
    # Calculate net electricity import and sum of generated electricity for
    # the specified countries
    electricity_consumption = calculate_net_electricity_import(
        snapshot_weightings,
        interesting_countries,
        EU27_countries,
        interesting_nodes,
        n_gross_elec,
    ) + sum_generated_electricity(
        snapshot_weightings,
        interesting_countries,
        EU27_countries,
        interesting_nodes,
        n_gross_elec,
    )

    electricity_consumption.name = "electricity_consumption"

    return electricity_consumption


def calculate_h2_gas_import(
    snapshot_weightings,
    interesting_countries,
    EU27_countries,
    interesting_nodes,
    network,
    carrier,
):
    """
    Calculates the net imports of methane or hydrogen for specified
    countries or regions by considering both inbound and outbound
    flows, including hydrogen/gas generated locally.

    Parameters
    ----------
    snapshot_weightings : pd.DataFrame
        DataFrame containing weightings for different snapshots in the
        analysis.
    interesting_countries : list of str
        List of country codes.
    EU27_countries : list of str
        List of EU27 country codes.
    interesting_nodes : list of str
        List of node identifiers.
    network : pypsa.Network
        The PyPSA Network object.
    carrier : str
        String of the carrier (gas, H2) for which the net imports
        should be calculated.

    Returns
    -------
    pd.Series
        A pandas Series containing the net imports of methane/hydrogen
        for each country or node. Small values near zero are set to
        zero to avoid misrepresentation due to rounding errors.
    """
    n_net_imports = network.copy()
    if carrier == "H2":
        # Define hydrogen transport mechanisms considered for imports
        import_carrier_list = [
            "H2 pipeline retrofitted",
            "H2 pipeline",
            "H2 pipeline (Kernnetz)",
            "import pipeline-H2",
            "import shipping-H2",
        ]
    elif carrier == "gas":
        # Define methane transport mechanisms considered for imports
        import_carrier_list = [
            "import lng gas",
            "import pipeline gas",
            "gas pipeline",
            "gas pipeline new",
            "gas pipeline tyndp",
            "import pipeline-syngas",
            "import shipping-syngas",
        ]

    # Dictionary to store hydrogen import sums for each identifier
    import_sums = {}  

    # Combine countries and nodes in one list for processing
    combined_identifiers = interesting_countries + interesting_nodes

    # Calculate net hydrogen imports for each interesting identifier
    for identifier in combined_identifiers:
        bus_slice, identifier_list = get_bus_slice(
            EU27_countries,
            identifier,
            interesting_countries,
            interesting_nodes,
        )

        # Determine indices for outgoing and incoming hydrogen/methane
        idx_pipes_out = n_net_imports.links[
            (n_net_imports.links.carrier.isin(import_carrier_list))
            & (n_net_imports.links.bus0.str[bus_slice].isin(identifier_list))
            & ~(n_net_imports.links.bus1.str[bus_slice].isin(identifier_list))
        ].index
        idx_pipes_in = n_net_imports.links[
            (n_net_imports.links.carrier.isin(import_carrier_list))
            & (n_net_imports.links.bus1.str[bus_slice].isin(identifier_list))
            & ~(n_net_imports.links.bus0.str[bus_slice].isin(identifier_list))
        ].index

        # Calculate exported and imported hydrogen/methane volumes
        e_pipes_out = (
            n_net_imports.links_t.p0[idx_pipes_out].mul(
                snapshot_weightings, axis=0
            )
        ).sum().sum() * 1e-6
        e_pipes_in = (
            -n_net_imports.links_t.p1[idx_pipes_in].mul(
                snapshot_weightings, axis=0
            )
        ).sum().sum() * 1e-6

        # Include hydrogen/methane generated within the identifier area
        idx_import_generators = n_net_imports.generators[
            (n_net_imports.generators.carrier.isin(import_carrier_list))
            & (
                n_net_imports.generators.bus.str[bus_slice].isin(
                    identifier_list
                )
            )
        ].index
        imp_import_generators = (
            n_net_imports.generators_t.p[idx_import_generators].mul(
                snapshot_weightings, axis=0
            )
        ).sum().sum() * 1e-6

        # Compute the net hydrogen import by summing the imports and
        # subtracting the exports
        import_sum = e_pipes_in - e_pipes_out + imp_import_generators
        import_sums[identifier] = (
            import_sum if import_sum >= 0.001 or import_sum <= -0.001 else 0
        )

    # Format the results as a pandas Series
    import_series = pd.Series(import_sums, name=f"net_import_{carrier}")
    import_series.index.name = "country_or_node"

    return import_series


def calculate_sum_generated_H2(
    snapshot_weightings,
    interesting_countries,
    EU27_countries,
    interesting_nodes,
    network,
):
    """
    Calculates the total hydrogen production for specified countries or
    regions by summing the outputs of hydrogen-related energy carriers.

    Parameters
    ----------
    snapshot_weightings : pd.DataFrame
        DataFrame containing weightings for different snapshots in the
        analysis.
    interesting_countries : list of str
        List of country codes.
    EU27_countries : list of str
        List of EU27 country codes.
    interesting_nodes : list of str
        List of node identifiers.
    network : pypsa.Network
        The PyPSA Network object.

    Returns
    -------
    pd.Series
        A pandas Series containing the total generated hydrogen for
        each country or node. Small values near zero are set to zero
        to avoid misrepresentation due to rounding errors.
    """
    n_gen_h2 = network.copy()
    generation_sums = {}

    # Specify carriers related to hydrogen production
    H2_gen_carrier = ["H2 Electrolysis", "SMR CC", "SMR"]

    # Combine countries and nodes in one list for processing
    combined_identifiers = interesting_countries + interesting_nodes

    # Calculate total hydrogen production for each specified identifier
    for identifier in combined_identifiers:
        bus_slice, identifier_list = get_bus_slice(
            EU27_countries,
            identifier,
            interesting_countries,
            interesting_nodes,
        )

        # Identify links associated with hydrogen production
        # and calculate generation
        idx_links = n_gen_h2.links[
            (n_gen_h2.links.carrier.isin(H2_gen_carrier))
            & (n_gen_h2.links.index.str[bus_slice].isin(identifier_list))
        ].index
        H2_generation = (
            (-n_gen_h2.links_t.p1[idx_links].mul(snapshot_weightings, axis=0))
            .sum()
            .fillna(0)
        ).sum() * 1e-6

        # Store hydrogen generation sum for the current identifier
        generation_sums[identifier] = (
            H2_generation if H2_generation >= 0.001 else 0
        )

    # Convert the generation sums into a pandas Series
    result_series = pd.Series(generation_sums, name="generated_hydrogen")
    result_series.index.name = "country_or_node"

    return result_series


def calculate_gross_hydrogen_consumption(
    snapshot_weightings,
    interesting_countries,
    EU27_countries,
    interesting_nodes,
    network,
):
    """
    Calculates the gross hydrogen consumption for specified countries
    or regions by adding the net hydrogen imports to the hydrogen
    generated within these regions.

    Parameters
    ----------
    snapshot_weightings : pd.DataFrame
        DataFrame containing weightings for different snapshots in the
        analysis.
    interesting_countries : list of str
        List of country codes.
    EU27_countries : list of str
        List of EU27 country codes.
    interesting_nodes : list of str
        List of node identifiers.
    network : pypsa.Network
        The PyPSA Network object.

    Returns
    -------
    pd.Series
        A pandas Series containing the gross hydrogen consumption for
        each country or node.
    """
    n_gross_h2 = network.copy()
    # Compute gross hydrogen consumption by adding net imports to generated
    # hydrogen
    hydrogen_consumption = calculate_h2_gas_import(
        snapshot_weightings,
        interesting_countries,
        EU27_countries,
        interesting_nodes,
        n_gross_h2,
        "H2",
    ) + calculate_sum_generated_H2(
        snapshot_weightings,
        interesting_countries,
        EU27_countries,
        interesting_nodes,
        n_gross_h2,
    )

    hydrogen_consumption.name = "hydrogen_consumption"

    return hydrogen_consumption


def calculate_capacity_thermal(
    interesting_countries, EU27_countries, interesting_nodes, network
):
    """
    Calculates thermal capacity sums for specified carriers grouped by
    fuel type across given countries or regions. This function maps
    various energy carriers to their respective fuel types and
    aggregates their optimized nominal power capacities.

    Parameters
    ----------
    interesting_countries : list of str
        List of country codes where the capacity is to be calculated.
    EU27_countries : list of str
        List of EU27 country codes, used specifically when the EU27
        group is selected.
    interesting_nodes : list of str
        List of node identifiers to include in the capacity
        calculation.
    network : pypsa.Network
        The PyPSA Network object, which contains all data about the
        energy system, including generators, links, and storage units.

    Returns
    -------
    pd.DataFrame
        DataFrame containing summed optimized nominal capacities (GW)
        for each fuel type, indexed by country or node.
    pd.DataFrame
        DataFrame containing total dispatch (TWh) for each fuel type,
        indexed by country or node.
    """
    n_cap_therm = network.copy()

    carrier_mapping = {
        "OCGT": "gas",
        "H2 Fuel Cell": "hydrogen",
        "H2 turbine": "hydrogen",
        "coal": "coal",
        "residential rural micro gas CHP": "gas",
        "services rural micro gas CHP": "gas",
        "residential urban decentral micro gas CHP": "gas",
        "services urban decentral micro gas CHP": "gas",
        "urban central gas CHP": "gas",
        "urban central gas CHP CC": "gas",
        "allam": "gas",
        "CCGT": "gas",
        "lignite": "lignite",
        "nuclear": "nuclear",
        "oil": "oil",
    }

    # Initialize dictionary to accumulate capacity sums grouped by fuel type
    capacity_sums = {group: {} for group in set(carrier_mapping.values())}

    # Initialize dictionary to accumulate dispatch sums grouped by fuel type
    dispatch_sums = {group: {} for group in set(carrier_mapping.values())}

    # Combine countries and nodes in one list for processing
    combined_identifiers = interesting_countries + interesting_nodes

    # Calculate capacity sums for each carrier and identifier
    for carrier, group in carrier_mapping.items():
        for identifier in combined_identifiers:
            bus_slice, identifier_list = get_bus_slice(
                EU27_countries,
                identifier,
                interesting_countries,
                interesting_nodes,
            )

            # Identify indices for relevant links and calculate their
            # optimized nominal power
            idx = n_cap_therm.links[
                (n_cap_therm.links.carrier == carrier)
                & (
                    n_cap_therm.links.index.str[bus_slice].isin(
                        identifier_list
                    )
                )
            ].index
            efficiency = n_cap_therm.links.loc[idx, "efficiency"]
            cap_carrier = (
                n_cap_therm.links.p_nom_opt[idx] * efficiency
            ).sum() * 1e-3

            # calculate total dispatch for this carrier in TWh
            dispatch_carrier = (
                -n_cap_therm.links_t.p1.loc[:, idx]
                .mul(n_cap_therm.snapshot_weightings.generators, axis=0)
                .sum()
                .sum()
                * 1e-6
            )
            # Add or update capacity and dispatch in the sum for the group
            # and identifier
            capacity_sums[group].setdefault(identifier, 0)
            capacity_sums[group][identifier] += cap_carrier
            dispatch_sums[group].setdefault(identifier, 0)
            dispatch_sums[group][identifier] += dispatch_carrier

    # Convert the capacity & dispatch sums into a Pandas DataFrame
    capacity_series = pd.DataFrame(capacity_sums)
    dispatch_series = pd.DataFrame(dispatch_sums)

    return capacity_series, dispatch_series


def calculate_capacity_of_H2_electrolysis(
    interesting_countries, EU27_countries, interesting_nodes, network
):
    """
    Calculates the total installed capacity of hydrogen electrolysis
    facilities for specified countries or regions. This includes the
    sum of all nominal capacities of links that have "H2 Electrolysis"
    in their carrier description.

    Parameters
    ----------
    interesting_countries : list of str
        List of country codes where the capacity is to be calculated.
    EU27_countries : list of str
        List of EU27 country codes, used specifically when the EU27
        group is selected.
    interesting_nodes : list of str
        List of node identifiers to include in the capacity
        calculation.
    network : pypsa.Network
        The PyPSA Network object, which contains all data about the
        energy system, including generators, links, and storage units.

    Returns
    -------
    pd.Series
        A pandas Series containing the total installed capacity of H2
        electrolysis for each country or node.
    """
    n_electrolysis = network.copy()

    capacity_sums = {}

    # Combine countries and nodes in one list for processing
    combined_identifiers = interesting_countries + interesting_nodes

    for identifier in combined_identifiers:
        bus_slice, identifier_list = get_bus_slice(
            EU27_countries,
            identifier,
            interesting_countries,
            interesting_nodes,
        )

        idx_H2 = n_electrolysis.links[
            n_electrolysis.links.carrier.str.contains("H2 Electrolysis")
            & n_electrolysis.links.bus0.str[bus_slice].isin(identifier_list)
        ].index

        # Calculate capacity for DC outputs
        cap_H2_electrolysis = (
            n_electrolysis.links.p_nom_opt[idx_H2].sum() * 1e-3
        )

        # Store the result in the dictionary
        capacity_sums[identifier] = cap_H2_electrolysis

    # Convert the dictionary to a Pandas Series
    capacity_series = pd.Series(capacity_sums, name="capacity_H2_electrolysis")
    capacity_series.index.name = "country_or_node"

    return capacity_series


def calculate_capacity_of_FT_synthesis(
    interesting_countries, EU27_countries, interesting_nodes, network
):
    """
    Calculates the total installed capacity of Fischer-Tropsch (FT)
    synthesis processes for specified countries or regions.
    This includes the sum of all nominal capacities of links that have
    "Fischer-Tropsch" in their carrier description.

    Parameters
    ----------
    interesting_countries : list of str
        List of country codes where the capacity is to be calculated.
    EU27_countries : list of str
        List of EU27 country codes, used specifically when the EU27
        group is selected.
    interesting_nodes : list of str
        List of node identifiers to include in the capacity
        calculation.
    network : pypsa.Network
        The PyPSA Network object, which contains all data about the
        energy system, including generators, links, and storage units.

    Returns
    -------
    pd.Series
        A pandas Series containing the total installed capacity of FT
        synthesis for each country or node.
    """
    n_cap_FTS = network.copy()
    capacity_sums = {}

    # Combine countries and nodes in one list for processing
    combined_identifiers = interesting_countries + interesting_nodes

    for identifier in combined_identifiers:
        bus_slice, identifier_list = get_bus_slice(
            EU27_countries,
            identifier,
            interesting_countries,
            interesting_nodes,
        )

        idx_FT = n_cap_FTS.links[
            n_cap_FTS.links.carrier.str.contains("Fischer-Tropsch")
            & n_cap_FTS.links.bus0.str[bus_slice].isin(identifier_list)
        ].index

        # Calculate capacity for FT synthesis outputs
        cap_FT_synthesis = n_cap_FTS.links.p_nom_opt[idx_FT].sum() * 1e-3

        # Store the result in the dictionary
        capacity_sums[identifier] = cap_FT_synthesis

    # Convert the dictionary to a Pandas Series
    capacity_series = pd.Series(capacity_sums, name="capacity_FT_synthesis")
    capacity_series.index.name = "country_or_node"

    return capacity_series


def calculate_generated_energy_H2_electrolysis(
    interesting_countries, EU27_countries, interesting_nodes, network
):
    """
    Calculates the energy generated from H2 Electrolysis links for
    specified countries or regions. This function utilizes energy
    balance statistics from the PyPSA network to identify and sum the
    energy outputs specific to H2 Electrolysis.

    Parameters
    ----------
    interesting_countries : list of str
        List of country codes where energy generation is to be
        calculated.
    EU27_countries : list of str
        List of EU27 country codes, used specifically when the EU27
        group is selected.
    interesting_nodes : list of str
        List of node identifiers to include in the energy calculation.
    network : pypsa.Network
        The PyPSA Network object, which contains all data about the
        energy system, including generators, links, and storage units.

    Returns
    -------
    pd.Series
        Series: A pandas Series containing the energy generated from H2
        Electrolysis for each country or node.
    """
    n_gen_electrol = network.copy()

    generated_energy = {}

    # Combine countries and nodes in one list for processing
    combined_identifiers = interesting_countries + interesting_nodes

    # Calculate generated energy from H2 Electrolysis
    # for each specified identifier
    for identifier in combined_identifiers:
        bus_slice, identifier_list = get_bus_slice(
            EU27_countries,
            identifier,
            interesting_countries,
            interesting_nodes,
        )

        idx_electrolysis = n_gen_electrol.links[
            (n_gen_electrol.links.carrier == "H2 Electrolysis")
            & n_gen_electrol.links.bus0.str[bus_slice].isin(identifier_list)
        ].index
        energy_demand = (
            -n_gen_electrol.links_t.p1.loc[:, idx_electrolysis]
            .mul(n_gen_electrol.snapshot_weightings.generators, axis=0)
            .sum()
            .sum()
            * 1e-6
        )

        # Store the generated energy for each identifier
        generated_energy[identifier] = energy_demand

    # Convert generated energy data into a pandas Series
    generated_energy_series = pd.Series(
        generated_energy, name="gen_energy_H2_electrolysis"
    )
    generated_energy_series.index.name = "country_or_node"

    return generated_energy_series


def calculate_generated_energy_FT_synthesis(
    snapshot_weightings,
    interesting_countries,
    EU27_countries,
    interesting_nodes,
    network,
):
    """
    Calculates the energy generated from Fischer-Tropsch synthesis
    processes for specified countries or regions. This function
    aggregates the outputs from links designated as 'Fischer-Tropsch'
    to compute the total energy produced.

    Parameters
    ----------
    snapshot_weightings : pd.DataFrame
        DataFrame containing weightings for different snapshots in the
        analysis.
    interesting_countries : list of str
        List of country codes where energy generation is to be
        calculated.
    EU27_countries : list of str
        List of EU27 country codes, used specifically when the EU27
        group is selected.
    interesting_nodes : list of str
        List of node identifiers to include in the energy calculation.
    network : pypsa.Network
        The PyPSA Network object, which contains all data about the
        energy system, including generators, links, and storage units.

    Returns
    -------
    pd.Series
        A pandas Series containing the energy generated from
        Fischer-Tropsch synthesis for each country or node.
    """
    n_gen_FTS = network.copy()

    generated_energy = {}

    # Combine countries and nodes in one list for processing
    combined_identifiers = interesting_countries + interesting_nodes

    # Calculate generated energy from Fischer-Tropsch synthesis
    # for each specified identifier
    for identifier in combined_identifiers:
        bus_slice, identifier_list = get_bus_slice(
            EU27_countries,
            identifier,
            interesting_countries,
            interesting_nodes,
        )

        # Filter links by 'Fischer-Tropsch' carrier
        # within the specified identifier
        idx_gen = n_gen_FTS.links[
            (n_gen_FTS.links.carrier == "Fischer-Tropsch")
            & (n_gen_FTS.links.index.str[bus_slice].isin(identifier_list))
        ].index

        # Calculate the total energy generated, adjusting units to TWh
        total_generation = -(
            (n_gen_FTS.links_t.p1[idx_gen].mul(snapshot_weightings, axis=0))
            .sum()
            .sum()
            * 1e-6
        )

        # Store the total energy generated for each identifier
        generated_energy[identifier] = total_generation

    # Convert the generated energy data into a pandas Series
    generated_energy_series = pd.Series(
        generated_energy, name="gen_energy_FT_synthesis"
    )
    generated_energy_series.index.name = "country_or_node"

    return generated_energy_series


def calculate_final_energy_demand_vector(
    snapshot_weightings,
    interesting_countries,
    EU27_countries,
    interesting_nodes,
    path_industrial_demand,
    mwh_coal_per_mwh_coke,
    network,
):
    """
    Calculates final energy demand across various sectors by processing
    energy balances and additional data. It considers specific energy
    carriers and adjusts for industrial demands, especially for coal
    and coke. The results are aggregated by energy groups and sectors,
    providing a detailed view of energy demands for key sectors like
    industry, transport, and households.

    Parameters
    ----------
    snapshot_weightings : pd.DataFrame
        DataFrame containing weightings for different snapshots in the
        analysis.
    interesting_countries : list of str
        List of country codes where energy demand is to be calculated.
    EU27_countries : list of str
        List of EU27 country codes, used specifically when the EU27
        group is selected.
    interesting_nodes : list of str
        List of node identifiers to include in the energy demand
        calculation.
    path_industrial_demand : str
        File path to the CSV containing specific industrial demand data
        for coal and coke.
    mwh_coal_per_mwh_coke : float
        Conversion factor from MWh of coke to MWh of coal used in
        industrial processes.
    network : pypsa.Network
        The PyPSA Network object, which contains all data about the
        energy system, including generators, links, and storage units.

    Returns
    -------
    pd.DataFrame
        A pandas DataFrame capturing total energy demands by energy
        groups and sectors, including a grand total.
        Each row corresponds to a country or node, with columns for
        each sector and energy type.
    """
    n_demand = network.copy()
    # Load industrial demand data
    print(path_industrial_demand)
    df_industrial_demand = pd.read_csv(path_industrial_demand)
    df_industrial_demand.rename(
        columns={df_industrial_demand.columns[0]: "bus"}, inplace=True
    )

    # Extract energy balance data for load and link components
    en_balance = n_demand.statistics.energy_balance(aggregate_bus=False)
    red_en_balance_load = en_balance.loc[
        en_balance.index.get_level_values(0).str.contains("Load")
    ]

    # need for adjustment: Aufschlüsselung der Sektoren nach Carriern
    # Define mappings for carriers to energy groups and sectors
    carrier_mapping = {
        "H2 for industry": "Hydrogen",
        "electricity": "Electrical Energy",
        "agriculture electricity": "Electrical Energy",
        "industry electricity": "Electrical Energy",
        "land transport EV": "Electrical Energy",
        "agriculture machinery oil": "Oil",
        "kerosene for aviation": "Oil",
        "land transport oil": "Oil",
        "naphtha for industry": "Oil",
        "shipping oil": "Oil",
        "coal for industry": "Coal",
        "gas for industry": "Gas",
        "shipping methanol": "Methanol",
        "solid biomass for industry": "Biomass",
        "land transport fuel cell": "Hydrogen",
        "services urban decentral heat": "Heat",
        "urban central heat": "Heat",
        "residential rural heat": "Heat",
        "residential urban decentral heat": "Heat",
        "services rural heat": "Heat",
    }

    sector_mapping = {
        "H2 for industry": "Industry",
        "industry electricity": "Industry",
        "naphtha for industry": "Industry",
        "coal for industry": "Industry",
        "gas for industry": "Industry",
        "solid biomass for industry": "Industry",
        "land transport EV": "Transport",
        "kerosene for aviation": "Transport",
        "land transport oil": "Transport",
        "shipping oil": "Transport",
        "land transport fuel cell": "Transport",
        "shipping methanol": "Transport",
        "electricity": "Households & Services",
        "urban central heat": "Households & Services",
        "agriculture electricity": "Agriculture",
        "services urban decentral heat": "Households & Services",
        "residential rural heat": "Households & Services",
        "residential urban decentral heat": "Households & Services",
        "services rural heat": "Households & Services",
        "agriculture machinery oil": "Agriculture",
    }

    # Initialize results containers for energy groups and sectors
    results = {group: {} for group in set(carrier_mapping.values())}
    sector_results = {sector: {} for sector in set(sector_mapping.values())}

    # Mean efficiency for BEV chargers for correct calculation of EV-loads
    eff_bev_charger = n_demand.links[
        n_demand.links.carrier == "BEV charger"
    ].efficiency.mean()

    # Combine countries and nodes in one list for processing
    combined_identifiers = interesting_countries + interesting_nodes

    # Calculate energy demand for each carrier and identifier
    for carrier, energy_group in carrier_mapping.items():
        for identifier in combined_identifiers:
            bus_slice, identifier_list = get_bus_slice(
                EU27_countries,
                identifier,
                interesting_countries,
                interesting_nodes,
            )

            # Energy demand calculation for heating carriers from load balance
            country_data = red_en_balance_load[
                red_en_balance_load.index.get_level_values(3)
                .str[bus_slice]
                .isin(identifier_list)
            ]
            carrier_data = country_data[
                country_data.index.get_level_values(1) == carrier
            ]
            energy_demand = -carrier_data.sum() * 1e-6

            # Additional calculations for coal using
            # industrial demand adjustments
            if energy_group == "Coal":
                df_index = df_industrial_demand[
                    df_industrial_demand.bus.str[bus_slice].isin(
                        identifier_list
                    )
                ].index
                industrial_demand_coal = (
                    df_industrial_demand.loc[df_index, "coal"].sum()
                    + mwh_coal_per_mwh_coke
                    * df_industrial_demand.loc[df_index, "coke"].sum()
                )
                energy_demand += industrial_demand_coal
            elif carrier == "land transport EV":
                idx = n_demand.loads[
                    (n_demand.loads.carrier == carrier)
                    & (n_demand.loads.bus.str[bus_slice].isin(identifier_list))
                ].index
                total_load = (
                    (
                        n_demand.loads_t.p[idx].mul(
                            snapshot_weightings, axis=0
                        )
                    ).sum()
                    * 1e-6
                    / eff_bev_charger
                )
                energy_demand = total_load.sum()

            # Accumulate energy demand into results containers
            sector = sector_mapping[carrier]
            results[energy_group].setdefault(identifier, 0)
            results[energy_group][identifier] += energy_demand
            sector_results[sector].setdefault(identifier, 0)
            sector_results[sector][identifier] += energy_demand

    # Create and combine DataFrames for results and sectors
    result_df = pd.DataFrame(results)
    sector_df = pd.DataFrame(sector_results)
    final_df = pd.concat([result_df, sector_df], axis=1)
    final_df["Total"] = final_df[
        ["Industry", "Transport", "Households & Services", "Agriculture"]
    ].sum(axis=1)

    return final_df


def calculate_capacity_of_electricity_interconnectors(
    interesting_countries, EU27_countries, interesting_nodes, network
):
    """
    Calculates the total installed capacity of both AC and DC
    electricity links and lines for specified countries or regions.
    This function assesses the capacities to provide a comprehensive
    view of cross-border electricity transmission
    capabilities.

    Parameters
    ----------
    interesting_countries : list of str
        List of country codes where the capacity is to be calculated.
    EU27_countries : list of str
        List of EU27 country codes, used specifically when the EU27
        group is selected.
    interesting_nodes : list of str
        List of node identifiers to include in the capacity
        calculation.
    network : pypsa.Network
        The PyPSA Network object, which contains all data about the
        energy system, including generators, links, and storage units.

    Returns
    -------
    pd.Series
        A pandas Series containing the total installed capacity of
        electricity links and lines for each country or node.
    """
    n_elec_inter = network.copy()
    # Dictionary to store the sums of capacities for each identifier
    capacity_sums = {}

    # Combine countries and nodes in one list for processing
    combined_identifiers = interesting_countries + interesting_nodes

    for identifier in combined_identifiers:
        bus_slice, identifier_list = get_bus_slice(
            EU27_countries,
            identifier,
            interesting_countries,
            interesting_nodes,
        )
        # Indices for DC outputs
        idx_DC_out = n_elec_inter.links[
            (n_elec_inter.links.carrier == "DC")
            & ~(n_elec_inter.links.index.str.contains("reversed"))
            & (n_elec_inter.links.bus0.str[bus_slice].isin(identifier_list))
            & ~(n_elec_inter.links.bus1.str[bus_slice].isin(identifier_list))
        ].index
        idx_DC_in = n_elec_inter.links[
            (n_elec_inter.links.carrier == "DC")
            & ~(n_elec_inter.links.index.str.contains("reversed"))
            & (n_elec_inter.links.bus1.str[bus_slice].isin(identifier_list))
            & ~(n_elec_inter.links.bus0.str[bus_slice].isin(identifier_list))
        ].index

        # Calculate capacity for DC outputs
        cap_DC_out = n_elec_inter.links.p_nom_opt.loc[idx_DC_out].sum() * 1e-3
        cap_DC_in = n_elec_inter.links.p_nom_opt.loc[idx_DC_in].sum() * 1e-3

        # Indices for AC outputs
        idx_AC_out = n_elec_inter.lines[
            (n_elec_inter.lines.bus0.str[bus_slice] == identifier)
            & ~(n_elec_inter.lines.bus1.str[bus_slice].isin(identifier_list))
        ].index
        idx_AC_in = n_elec_inter.lines[
            (n_elec_inter.lines.bus1.str[bus_slice] == identifier)
            & ~(n_elec_inter.lines.bus0.str[bus_slice].isin(identifier_list))
        ].index

        # Calculate capacity for AC outputs
        cap_AC_out = n_elec_inter.lines.s_nom_opt[idx_AC_out].sum() * 1e-3
        cap_AC_in = n_elec_inter.lines.s_nom_opt[idx_AC_in].sum() * 1e-3

        # Total capacity sum for the current identifier
        total_capacity = cap_DC_out + cap_DC_in + cap_AC_out + cap_AC_in

        # Store the result in the dictionary
        capacity_sums[identifier] = total_capacity

    # Convert the dictionary to a Pandas Series
    capacity_series = pd.Series(
        capacity_sums, name="capacity_Electricity_interconnectors"
    )
    capacity_series.index.name = "country_or_node"

    return capacity_series


def calculate_capacity_of_electricity_links_and_lines(
    interesting_countries, EU27_countries, interesting_nodes, network
):
    """
    Calculates the total installed capacity of both AC and DC
    electricity links and lines for specified countries or regions.
    This function assesses the capacities to provide a comprehensive
    view of internal and cross-border electricity transmission
    capabilities.

    Parameters
    ----------
    interesting_countries : list of str
        List of country codes where the capacity is to be calculated.
    EU27_countries : list of str
        List of EU27 country codes, used specifically when the EU27
        group is selected.
    interesting_nodes : list of str
        List of node identifiers to include in the capacity
        calculation.
    network : pypsa.Network
        The PyPSA Network object, which contains all data about the
        energy system, including generators, links, and storage units.

    Returns
    -------
    pd.Series
        A pandas Series containing the total installed capacity of
        electricity links and lines for each country or node.
    """
    n_elec_trans = network.copy()
    # Dictionary to store the sums of capacities for each identifier
    capacity_sums = {}

    # Combine countries and nodes in one list for processing
    combined_identifiers = interesting_countries + interesting_nodes

    for identifier in combined_identifiers:
        bus_slice, identifier_list = get_bus_slice(
            EU27_countries,
            identifier,
            interesting_countries,
            interesting_nodes,
        )
        # Indices for DC links
        idx_DC_out = n_elec_trans.links[
            (n_elec_trans.links.carrier == "DC")
            & ~(n_elec_trans.links.index.str.contains("reversed"))
            & (n_elec_trans.links.bus0.str[bus_slice].isin(identifier_list))
            & ~(n_elec_trans.links.bus1.str[bus_slice].isin(identifier_list))
        ].index
        idx_DC_in = n_elec_trans.links[
            (n_elec_trans.links.carrier == "DC")
            & ~(n_elec_trans.links.index.str.contains("reversed"))
            & (n_elec_trans.links.bus1.str[bus_slice].isin(identifier_list))
            & ~(n_elec_trans.links.bus0.str[bus_slice].isin(identifier_list))
        ].index
        idx_DC_within = n_elec_trans.links[
            (n_elec_trans.links.carrier == "DC")
            & ~(n_elec_trans.links.index.str.contains("reversed"))
            & (n_elec_trans.links.bus0.str[bus_slice].isin(identifier_list))
            & (n_elec_trans.links.bus1.str[bus_slice].isin(identifier_list))
        ].index

        # Calculate capacity for DC links
        cap_DC_GWkm_out = (
            n_elec_trans.links.p_nom_opt.loc[idx_DC_out]
            * n_elec_trans.links.length.loc[idx_DC_out]
        ).sum() * 1e-3
        cap_DC_GWkm_in = (
            n_elec_trans.links.p_nom_opt.loc[idx_DC_in]
            * n_elec_trans.links.length.loc[idx_DC_in]
        ).sum() * 1e-3
        cap_DC_GWkm_within = (
            n_elec_trans.links.p_nom_opt.loc[idx_DC_within]
            * n_elec_trans.links.length.loc[idx_DC_within]
        ).sum() * 1e-3

        # Indices for AC links
        idx_AC_out = n_elec_trans.lines[
            (n_elec_trans.lines.bus0.str[bus_slice].isin(identifier_list))
            & ~(n_elec_trans.lines.bus1.str[bus_slice].isin(identifier_list))
        ].index
        idx_AC_in = n_elec_trans.lines[
            (n_elec_trans.lines.bus1.str[bus_slice].isin(identifier_list))
            & ~(n_elec_trans.lines.bus0.str[bus_slice].isin(identifier_list))
        ].index
        idx_AC_within = n_elec_trans.lines[
            (n_elec_trans.lines.bus0.str[bus_slice].isin(identifier_list))
            & (n_elec_trans.lines.bus1.str[bus_slice].isin(identifier_list))
        ].index

        # Calculate capacity for AC links
        cap_AC_out = (
            n_elec_trans.lines.s_nom_opt.loc[idx_AC_out]
            * n_elec_trans.lines.length[idx_AC_out]
        ).sum() * 1e-3
        cap_AC_in = (
            n_elec_trans.lines.s_nom_opt.loc[idx_AC_in]
            * n_elec_trans.lines.length[idx_AC_in]
        ).sum() * 1e-3
        cap_AC_within = (
            n_elec_trans.lines.s_nom_opt.loc[idx_AC_within]
            * n_elec_trans.lines.length[idx_AC_within]
        ).sum() * 1e-3

        # Total capacity sum for the current identifier
        total_capacity = (
            cap_DC_GWkm_out
            + cap_DC_GWkm_in
            + cap_DC_GWkm_within
            + cap_AC_out
            + cap_AC_in
            + cap_AC_within
        )

        # Store the result in the dictionary
        capacity_sums[identifier] = total_capacity

    # Convert the dictionary to a Pandas Series
    capacity_series = pd.Series(
        capacity_sums, name="capacity_Electricity_links_and_lines"
    )
    capacity_series.index.name = "country_or_node"

    return capacity_series


def calculate_capacity_interconnectors(
    interesting_countries, EU27_countries, interesting_nodes, network, carrier
):
    """
    Calculates the total installed capacity of pipeline interconnectors
    for specified countries or regions. This function assesses the
    capacities of transport via pipelines, reflecting both inbound and
    outbound capabilities. The medium of transport (gas, H2, co2) is
    determined by carrier.

    Parameters
    ----------
    interesting_countries : list of str
        List of country codes where the capacity is to be calculated.
    EU27_countries : list of str
        List of EU27 country codes, used specifically when the EU27
        group is selected.
    interesting_nodes : list of str
        List of node identifiers to include in the capacity
        calculation.
    network : pypsa.Network
        The PyPSA Network object, which contains all data about the
        energy system, including generators, links, and storage units.
    carrier : str
        Selection of 'H2', 'gas', 'co2' to determine the type of
        pipeline.

    Returns
    -------
    pd.Series
        A pandas Series containing the total installed capacity of
        interconnectors for each country or node.
    """
    n_intercon = network.copy()

    # Dictionary to store the sums of capacities for each identifier
    capacity_sums = {}

    # Combine countries and nodes in one list for processing
    combined_identifiers = interesting_countries + interesting_nodes

    for identifier in combined_identifiers:
        bus_slice, identifier_list = get_bus_slice(
            EU27_countries,
            identifier,
            interesting_countries,
            interesting_nodes,
        )
        # Indices for pipeline interconnectors
        idx_pipe_out = n_intercon.links[
            (n_intercon.links.carrier.str.contains(f"{carrier} pipeline"))
            & ~(n_intercon.links.index.str.contains("reversed"))
            & (n_intercon.links.bus0.str[bus_slice].isin(identifier_list))
            & ~(n_intercon.links.bus1.str[bus_slice].isin(identifier_list))
        ].index
        idx_pipe_in = n_intercon.links[
            (n_intercon.links.carrier.str.contains(f"{carrier} pipeline"))
            & ~(n_intercon.links.index.str.contains("reversed"))
            & (n_intercon.links.bus1.str[bus_slice].isin(identifier_list))
            & ~(n_intercon.links.bus0.str[bus_slice].isin(identifier_list))
        ].index

        # Calculate capacity for pipeline interconnectors
        cap_pipe_out = (
            n_intercon.links.p_nom_opt.loc[idx_pipe_out].sum() * 1e-3
        )
        cap_pipe_in = n_intercon.links.p_nom_opt.loc[idx_pipe_in].sum() * 1e-3

        # Total capacity sum for the current identifier
        total_intercon_capacity = cap_pipe_out + cap_pipe_in

        # Store the result in the dictionary
        capacity_sums[identifier] = total_intercon_capacity

    # Convert the dictionary to a Pandas Series
    capacity_series = pd.Series(
        capacity_sums, name=f"capacity_{carrier}_interconnectors"
    )
    capacity_series.index.name = "country_or_node"

    return capacity_series


def calculate_capacity_grids(
    interesting_countries, EU27_countries, interesting_nodes, network, carrier
):
    """
    Calculates the total installed capacity of pipelines for specified
    countries or regions. This function assesses the capacities of
    pipelines, accounting for both outbound, inbound, and
    within-country transfer capabilities. The medium of transport is
    given by carrier.

    Parameters
    ----------
    interesting_countries : list of str
        List of country codes where the capacity is to be calculated.
    EU27_countries : list of str
        List of EU27 country codes, used specifically when the EU27
        group is selected.
    interesting_nodes : list of str
        List of node identifiers to include in the capacity
        calculation.
    network : pypsa.Network
        The PyPSA Network object, which contains all data about the
        energy system, including generators, links, and storage units.
    carrier : str
        Type of pipeline to consider ('H2', 'gas', 'co2').

    Returns
    -------
    pd.Series
        A pandas Series containing the total installed capacity of
        pipeline network for each country or node.
    """
    n_grids = network.copy()

    # Dictionary to store the sums of capacities for each identifier
    capacity_sums = {}

    # Combine countries and nodes in one list for processing
    combined_identifiers = interesting_countries + interesting_nodes

    for identifier in combined_identifiers:
        bus_slice, identifier_list = get_bus_slice(
            EU27_countries,
            identifier,
            interesting_countries,
            interesting_nodes,
        )
        # Indices for pipeline links
        idx_pipes_out = n_grids.links[
            (n_grids.links.carrier.str.contains(f"{carrier} pipeline"))
            & ~(n_grids.links.index.str.contains("reversed"))
            & (n_grids.links.bus0.str[bus_slice].isin(identifier_list))
            & ~(n_grids.links.bus1.str[bus_slice].isin(identifier_list))
        ].index
        idx_pipes_in = n_grids.links[
            (n_grids.links.carrier.str.contains(f"{carrier} pipeline"))
            & ~(n_grids.links.index.str.contains("reversed"))
            & (n_grids.links.bus1.str[bus_slice].isin(identifier_list))
            & ~(n_grids.links.bus0.str[bus_slice].isin(identifier_list))
        ].index
        idx_pipes_within = n_grids.links[
            (n_grids.links.carrier.str.contains(f"{carrier} pipeline"))
            & ~(n_grids.links.index.str.contains("reversed"))
            & (n_grids.links.bus0.str[bus_slice].isin(identifier_list))
            & (n_grids.links.bus1.str[bus_slice].isin(identifier_list))
        ].index

        # Calculate capacity for pipeline links
        cap_pipes_GWkm_out = (
            n_grids.links.p_nom_opt.loc[idx_pipes_out]
            * n_grids.links.length.loc[idx_pipes_out]
        ).sum() * 1e-3
        cap_pipes_GWkm_in = (
            n_grids.links.p_nom_opt.loc[idx_pipes_in]
            * n_grids.links.length.loc[idx_pipes_in]
        ).sum() * 1e-3
        cap_pipes_GWkm_within = (
            n_grids.links.p_nom_opt.loc[idx_pipes_within]
            * n_grids.links.length.loc[idx_pipes_within]
        ).sum() * 1e-3

        # Total capacity sum for the current identifier
        total_pipes_capacity = (
            cap_pipes_GWkm_out + cap_pipes_GWkm_in + cap_pipes_GWkm_within
        )

        # Store the result in the dictionary
        capacity_sums[identifier] = total_pipes_capacity

    # Convert the dictionary to a Pandas Series for
    # better visualization and usage
    capacity_series = pd.Series(
        capacity_sums, name=f"capacity_{carrier}_pipelines"
    )
    capacity_series.index.name = "country_or_node"

    return capacity_series


def calculate_primary_energy_demand_gas(
    snapshot_weightings,
    interesting_countries,
    EU27_countries,
    interesting_nodes,
    network,
):
    """
    Calculates the primary energy demand for methane for specified
    countries or regions by accounting for both inbound and outbound
    flows as well as production within the regions. This includes
    assessing the net balance of natural gas and synthetic gas (syngas)
    through pipelines and other import channels.

    Parameters
    ----------
    snapshot_weightings : pd.DataFrame
        DataFrame containing weightings for different snapshots in the
        analysis.
    interesting_countries : list of str
        List of country codes where the net imports are to be
        calculated.
    EU27_countries : list of str
        List of EU27 country codes, used specifically when the EU27
        group is selected.
    interesting_nodes : list of str
        List of node identifiers to include in the net import
        calculation.
    network : pypsa.Network
        The PyPSA Network object, which contains all data about the
        energy system, including generators, links, and storage units.

    Returns
    -------
    pd.Series
        tuple: Two pandas Series objects:
            - The first Series contains the net imports of natural gas
              for each country or node.
            - The second Series contains the net imports of synthetic
              gas for each country or node.
    """
    n_gas_import = network.copy()
    # Definitions of carriers for gas and syngas imports
    natural_gas_carriers = [
        "import lng gas",
        "import pipeline gas",
        "production gas",
        "gas pipeline",
        "gas pipeline new",
        "gas pipeline tyndp",
    ]
    syngas_carriers = ["biogas to gas", "biogas to gas CC", "syngas to gas"]

    gas_demand = {}

    # Combine countries and nodes in one list for processing
    combined_identifiers = interesting_countries + interesting_nodes

    for identifier in combined_identifiers:
        bus_slice, identifier_list = get_bus_slice(
            EU27_countries,
            identifier,
            interesting_countries,
            interesting_nodes,
        )

        # Identify and sum outbound and inbound gas flows
        idx_gas_out = n_gas_import.links[
            (n_gas_import.links.carrier.isin(natural_gas_carriers))
            & (n_gas_import.links.bus0.str[bus_slice].isin(identifier_list))
            & ~(n_gas_import.links.bus1.str[bus_slice].isin(identifier_list))
        ].index
        idx_gas_in = n_gas_import.links[
            (n_gas_import.links.carrier.isin(natural_gas_carriers))
            & (n_gas_import.links.bus1.str[bus_slice].isin(identifier_list))
            & ~(n_gas_import.links.bus0.str[bus_slice].isin(identifier_list))
        ].index
        imp_gas_out = (
            -(
                n_gas_import.links_t.p0[idx_gas_out].mul(
                    snapshot_weightings, axis=0
                )
            )
            .sum()
            .sum()
            * 1e-6
        )
        imp_gas_in = (
            -(
                n_gas_import.links_t.p1[idx_gas_in].mul(
                    snapshot_weightings, axis=0
                )
            )
            .sum()
            .sum()
            * 1e-6
        )

        # Include natural gas imported and generated within the country
        idx_gas_generators = n_gas_import.generators[
            (n_gas_import.generators.carrier.isin(natural_gas_carriers))
            & (n_gas_import.generators.bus.str[:2].isin(identifier_list))
        ].index
        imp_gas_generators = (
            n_gas_import.generators_t.p[idx_gas_generators].mul(
                snapshot_weightings, axis=0
            )
        ).sum().sum() * 1e-6

        # Identify links processing biomethane and synthetic gas
        # and calculate total energy converted
        en_link_syngas_index = n_gas_import.links[
            (n_gas_import.links.carrier.isin(syngas_carriers))
            & (n_gas_import.links.bus1.str[bus_slice].isin(identifier_list))
        ].index
        en_link_syngas = abs(
            (
                n_gas_import.links_t.p1[en_link_syngas_index].mul(
                    snapshot_weightings, axis=0
                )
            )
            .sum()
            .sum()
            * 1e-6
        )

        # Compute the net import for gas
        natural_gas_demand = imp_gas_in + imp_gas_out + imp_gas_generators

        # Put together gas, syngas and biogas
        gas_demand[identifier] = natural_gas_demand + en_link_syngas

    # Convert import sums to pandas Series
    gas_import_series = pd.Series(gas_demand, name="gas_primary_energy_demand")

    return gas_import_series


def calculate_primary_energy_demand_oil(
    snapshot_weightings,
    interesting_countries,
    EU27_countries,
    interesting_nodes,
    network,
):
    """
    Calculates the primary energy demand from oil and oil-related
    processes for specified countries or regions. This function
    assesses energy consumption related to traditional oil use,
    Fischer-Tropsch synthesis, and biomass-to-liquid processes.

    Parameters
    ----------
    snapshot_weightings : pd.DataFrame
        DataFrame containing weightings for different snapshots in the
        analysis.
    interesting_countries : list of str
        List of country codes where the energy demand is to be
        calculated.
    EU27_countries : list of str
        List of EU27 country codes, used specifically when the EU27
        group is selected.
    interesting_nodes : list of str
        List of node identifiers to include in the energy demand
        calculation.
    network : pypsa.Network
        The PyPSA Network object, which contains all data about the
        energy system, including generators, links, and storage units.

    Returns
    -------
    pd.Series
        A pandas Series containing the primary energy demand from oil
        for each country or node.
    """
    n_prim_oil = network.copy()

    oil_energy_demand_sums = {}

    # Combine countries and nodes in one list for processing
    combined_identifiers = interesting_countries + interesting_nodes

    for identifier in combined_identifiers:
        bus_slice, identifier_list = get_bus_slice(
            EU27_countries,
            identifier,
            interesting_countries,
            interesting_nodes,
        )

        # Calculate energy demand for Fischer-Tropsch synthesis
        index_fischer_tropsch = n_prim_oil.links[
            (n_prim_oil.links.carrier.str.contains("Fischer-Tropsch"))
            & (n_prim_oil.links.bus0.str[bus_slice].isin(identifier_list))
        ].index
        fischer_tropsch_synthesis = (
            n_prim_oil.links_t.p1[index_fischer_tropsch].mul(
                snapshot_weightings, axis=0
            )
        ).sum().sum() * 1e-6

        # Calculate energy demand for biomass to liquid processes
        index_biomass_to_liquid = n_prim_oil.links[
            (n_prim_oil.links.carrier == "biomass to liquid")
            & (n_prim_oil.links.bus0.str[bus_slice].isin(identifier_list))
        ].index
        en_biomass_to_liquid = (
            n_prim_oil.links_t.p1[index_biomass_to_liquid].mul(
                snapshot_weightings, axis=0
            )
        ).sum().sum() * 1e-6

        # Calculate energy for oil-related carriers
        # using generator and link data
        oil_carriers = [
            "land transport oil",
            "shipping oil",
            "residential rural oil boiler",
            "services rural oil boiler",
            "residential urban decentral oil boiler",
            "services urban decentral oil boiler",
            "naphtha for industry",
            "kerosene for aviation",
            "agriculture machinery oil",
            "oil",
        ]

        oil_demand_idx = n_prim_oil.links[
            (n_prim_oil.links.carrier.isin(oil_carriers))
            & (n_prim_oil.links.bus1.str[bus_slice].isin(identifier_list))
        ].index
        link_oil_demand = (
            n_prim_oil.links_t.p0[oil_demand_idx].mul(
                snapshot_weightings, axis=0
            )
        ).sum().sum() * 1e-6

        # Total energy demand calculation for the identifier
        total_energy = (
            fischer_tropsch_synthesis + link_oil_demand + en_biomass_to_liquid
        )
        oil_energy_demand_sums[identifier] = total_energy

    # Convert the oil energy demand sums into a pandas Series
    oil_energy_demand_series = pd.Series(
        oil_energy_demand_sums, name="primary_energy_oil"
    )
    oil_energy_demand_series.index.name = "country_or_node"

    return oil_energy_demand_series


def calculate_primary_energy_demand_coal(
    snapshot_weightings,
    interesting_countries,
    EU27_countries,
    interesting_nodes,
    path_industrial_demand,
    mwh_coal_per_mwh_coke,
    network,
):
    """
    Calculates the primary energy demand from coal and coke for
    specified countries or regions. This includes the sum of coal
    energy transmitted through the network and the coal used directly
    in industries, adjusted for coke energy content.

    Parameters
    ----------
    snapshot_weightings : pd.DataFrame
        DataFrame containing weightings for different snapshots in the
        analysis.
    interesting_countries : list of str
        List of country codes where the energy demand is to be
        calculated.
    EU27_countries : list of str
        List of EU27 country codes, used specifically when the EU27
        group is selected.
    interesting_nodes : list of str
        List of node identifiers to include in the energy demand
        calculation.
    path_industrial_demand : str
        File path to the CSV containing specific industrial demand data
        for coal and coke.
    mwh_coal_per_mwh_coke : float
        Conversion factor from MWh of coke to MWh of coal used in
        industrial processes.
    network : pypsa.Network
        The PyPSA Network object, which contains all data about the
        energy system, including generators, links, and storage units.

    Returns
    -------
    pd.Series
        A pandas Series containing the primary energy demand from coal
        for each country or node.
    """
    n_coal = network.copy()
    df_industrial_demand = pd.read_csv(path_industrial_demand)
    df_industrial_demand.rename(
        columns={df_industrial_demand.columns[0]: "bus"}, inplace=True
    )

    energy_sums = {}  # Dictionary to store the energy sums for each identifier

    # Combine countries and nodes in one list for processing
    combined_identifiers = interesting_countries + interesting_nodes

    for identifier in combined_identifiers:
        bus_slice, identifier_list = get_bus_slice(
            EU27_countries,
            identifier,
            interesting_countries,
            interesting_nodes,
        )

        # Compute energy transfer from coal-related links
        # filtered by identifier
        link_index = n_coal.links[
            (n_coal.links.carrier == "coal")
            & (n_coal.links.bus1.str[bus_slice].isin(identifier_list))
        ].index
        link_coal_energy = (
            n_coal.links_t.p0[link_index].mul(snapshot_weightings, axis=0)
        ).sum().sum() * 1e-6

        # Aggregate coal and coke industrial demands adjusted
        # for energy content
        df_index = df_industrial_demand[
            df_industrial_demand.bus.str[bus_slice].isin(identifier_list)
        ].index
        industrial_demand_coal = (
            df_industrial_demand.coal[df_index].sum()
            + mwh_coal_per_mwh_coke * df_industrial_demand.coke[df_index].sum()
        )

        # Sum up the energy demand from transmission links and industrial uses
        energy_sums[identifier] = link_coal_energy + industrial_demand_coal

    # Format the results as a Pandas Series
    energy_demand_series = pd.Series(
        energy_sums, name="primary_energy_demand_coal"
    )
    energy_demand_series.index.name = "country_or_node"

    return energy_demand_series


def calculate_primary_energy_demand_lignite(
    snapshot_weightings,
    interesting_countries,
    EU27_countries,
    interesting_nodes,
    network,
):
    """
    Calculates the primary energy demand from lignite for specified
    countries or regions. This function assesses the total energy
    associated with lignite in the network.

    Parameters
    ----------
    snapshot_weightings : pd.DataFrame
        DataFrame containing weightings for different snapshots in the
        analysis.
    interesting_countries : list of str
        List of country codes where the energy demand is to be
        calculated.
    EU27_countries : list of str
        List of EU27 country codes, used specifically when the EU27
        group is selected.
    interesting_nodes : list of str
        List of node identifiers to include in the energy demand
        calculation.
    network : pypsa.Network
        The PyPSA Network object, which contains all data about the
        energy system, including generators, links, and storage units.

    Returns
    -------
    pd.Series
        A pandas Series containing the primary energy demand from
        lignite for each country or node.
    """
    n_lignite = network.copy()
    energy_sums = {}  # Dictionary to store the energy sums for each identifier

    # Combine countries and nodes in one list for processing
    combined_identifiers = interesting_countries + interesting_nodes

    for identifier in combined_identifiers:
        bus_slice, identifier_list = get_bus_slice(
            EU27_countries,
            identifier,
            interesting_countries,
            interesting_nodes,
        )

        # Identify links associated with lignite and calculate total energy
        link_index = n_lignite.links[
            (n_lignite.links.carrier == "lignite")
            & (n_lignite.links.bus1.str[bus_slice].isin(identifier_list))
        ].index
        link_energy = (
            n_lignite.links_t.p0[link_index].mul(snapshot_weightings, axis=0)
        ).sum().sum() * 1e-6

        # Store the lignite energy demand for the identifier
        energy_sums[identifier] = link_energy

    # Convert the dictionary of energy sums to a Pandas Series
    energy_demand_series = pd.Series(
        energy_sums, name="primary_energy_demand_lignite"
    )
    energy_demand_series.index.name = "country_or_node"

    return energy_demand_series


def calculate_primary_energy_demand_nuclear(
    snapshot_weightings,
    interesting_countries,
    EU27_countries,
    interesting_nodes,
    network,
):
    """
    Calculates the primary energy demand from nuclear power for
    specified countries or regions. This function assesses the total
    energy generated by nuclear power stations in
    the network.

    Parameters
    ----------
    snapshot_weightings : pd.DataFrame
        DataFrame containing weightings for different snapshots in the
        analysis.
    interesting_countries : list of str
        List of country codes where the energy demand is to be
        calculated.
    EU27_countries : list of str
        List of EU27 country codes, used specifically when the EU27
        group is selected.
    interesting_nodes : list of str
        List of node identifiers to include in the energy demand
        calculation.
    network : pypsa.Network
        The PyPSA Network object, which contains all data about the
        energy system, including generators, links, and storage units.

    Returns
    -------
    pd.Series
        A pandas Series containing the primary energy demand from
        nuclear power for each country or node.
    """
    n_nuclear = network.copy()

    energy_sums = {}  # Dictionary to store the energy sums for each identifier

    # Combine countries and nodes in one list for processing
    combined_identifiers = interesting_countries + interesting_nodes

    for identifier in combined_identifiers:
        bus_slice, identifier_list = get_bus_slice(
            EU27_countries,
            identifier,
            interesting_countries,
            interesting_nodes,
        )
        # Filter nuclear power links by identifier
        # and calculate total energy transferred
        link_index = n_nuclear.links[
            (n_nuclear.links.carrier == "nuclear")
            & (n_nuclear.links.bus1.str[bus_slice].isin(identifier_list))
        ].index
        link_nuclear_energy = (
            n_nuclear.links_t.p0[link_index].mul(snapshot_weightings, axis=0)
        ).sum().sum() * 1e-6

        # Store the computed nuclear energy demand for the identifier
        energy_sums[identifier] = link_nuclear_energy

    # Format the results as a Pandas Series
    energy_demand_series = pd.Series(
        energy_sums, name="primary_energy_demand_nuclear"
    )
    energy_demand_series.index.name = "country_or_node"

    return energy_demand_series


def calculate_primary_energy_demand_biomass(
    snapshot_weightings,
    interesting_countries,
    EU27_countries,
    interesting_nodes,
    network,
):
    """
    Calculates the primary energy demand from biomass for specified
    countries or regions. This function assesses the total energy
    associated with biomass storage facilities and sums up the stored
    energy across the network.

    Parameters
    ----------
    snapshot_weightings : DataFrame
        DataFrame containing weightings for different snapshots in the
        analysis.
    interesting_countries : list of str
        List of country codes where the energy demand is to be
        calculated.
    EU27_countries : list of str
        List of EU27 country codes, used specifically when the EU27
        group is selected.
    interesting_nodes : list of str
        List of node identifiers to include in the energy demand
        calculation.
    network : pypsa.Network
        The PyPSA Network object, which contains all data about the
        energy system, including generators, links, and storage units.

    Returns
    -------
    pd.Series
        A pandas Series containing the primary energy demand from
        biomass for each country or node.
    """
    n_biomass = network.copy()

    energy_sums = {}  # Dictionary to store the energy sums for each identifier

    # Combine countries and nodes in one list for processing
    combined_identifiers = interesting_countries + interesting_nodes

    for identifier in combined_identifiers:
        bus_slice, identifier_list = get_bus_slice(
            EU27_countries,
            identifier,
            interesting_countries,
            interesting_nodes,
        )
        # Identify biomass storage facilities relevant to the identifier
        # and sum stored energy
        en_store_biomass_index = n_biomass.stores[
            (n_biomass.stores.carrier == "solid biomass")
            & (n_biomass.stores.bus.str[bus_slice].isin(identifier_list))
        ].index
        en_store_biomass = (
            n_biomass.stores_t.p[en_store_biomass_index].mul(
                snapshot_weightings, axis=0
            )
        ).sum().sum() * 1e-6

        # Store the computed biomass energy demand for the identifier
        energy_sums[identifier] = en_store_biomass

    # Convert the results to a Pandas Series
    energy_demand_series = pd.Series(
        energy_sums, name="primary_energy_demand_biomass"
    )
    energy_demand_series.index.name = "country_or_node"

    return energy_demand_series


def calculate_primary_energy_demand_vector(
    snapshot_weightings,
    interesting_countries,
    EU27_countries,
    interesting_nodes,
    path_industrial_demand,
    mwh_coal_per_mwh_coke,
    network,
):
    """
    Computes the primary energy demand across multiple energy sources
    for specified countries and nodes. The function maps specific
    renewable generator carriers to more general categories and
    calculates the total generation from these for each country or node
    of interest. It also calculates various imports and demands for
    other energy types such as electricity, hydrogen, gas, oil, coal,
    lignite, nuclear and biomass. The total primary energy demand is
    then aggregated into a final DataFrame which includes detailed
    columns for each type of energy.

    Parameters
    ----------
    snapshot_weightings : pd.DataFrame
        DataFrame containing weightings for different snapshots in the
        analysis.
    interesting_countries : list of str
        List of country codes where the energy demand is to be
        calculated.
    EU27_countries : list of str
        List of EU27 country codes, used specifically when the EU27
        group is selected.
    interesting_nodes : list of str
        List of node identifiers to include in the energy demand
        calculation.
    path_industrial_demand : str
        File path to the CSV containing specific industrial demand data
        for coal and coke.
    mwh_coal_per_mwh_coke : float
        Conversion factor from MWh of coke to MWh of coal used in
        industrial processes.
    network : pypsa.Network
        The PyPSA Network object, which contains all data about the
        energy system, including generators, links, and storage units.

    Returns
    -------
    pd.DataFrame
        A comprehensive DataFrame capturing the total primary energy
        demand across all specified energy sources for each country or
        node, including detailed energy type breakdowns.
    """
    n_prim_vec = network.copy()
    # Map specific generator carriers to more general categories
    carrier_map = {
        "solar rooftop": "solar",
        "offwind-ac": "wind",
        "offwind-dc": "wind",
        "onwind": "wind",
        "ror": "water",
        "hydro": "water",
    }
    n_prim_vec.generators.carrier.replace(carrier_map, inplace=True)
    n_prim_vec.storage_units.carrier.replace(carrier_map, inplace=True)

    # List of interesting carriers to consider for the
    # energy demand calculation for renewables
    interesting_carriers = ["solar", "wind", "water"]

    # Initialize results dictionary to store energy sums
    # for each country and carrier
    results = {}

    # Combine countries and nodes in one list for processing
    combined_identifiers = interesting_countries + interesting_nodes

    # Iterate over each carrier of interest for calculating the
    # primary energy demand for renewable energies
    for carrier in interesting_carriers:
        carrier_generators = n_prim_vec.generators[
            n_prim_vec.generators.carrier == carrier
        ]
        carrier_storage_units = n_prim_vec.storage_units[
            n_prim_vec.storage_units.carrier == carrier
        ]
        total_generation_sum = {}

        # Iterate over each identifier of interest
        for identifier in combined_identifiers:
            bus_slice, identifier_list = get_bus_slice(
                EU27_countries,
                identifier,
                interesting_countries,
                interesting_nodes,
            )

            # Filter generators & storage units based on the identifier list
            idx_gen = carrier_generators[
                carrier_generators.bus.str[bus_slice].isin(identifier_list)
            ].index
            idx_su = carrier_storage_units[
                carrier_storage_units.bus.str[bus_slice].isin(identifier_list)
            ].index

            # Calculate the total generation for the selected generators
            # and resolution
            total_generation = (
                n_prim_vec.generators_t.p[idx_gen].mul(
                    snapshot_weightings, axis=0
                )
            ).sum().sum() * 1e-6

            # Calculate the total generation from storage units
            # for the selected generators and resolution
            total_generation_su = (
                n_prim_vec.storage_units_t.p[idx_su].mul(
                    snapshot_weightings, axis=0
                )
            ).sum().sum() * 1e-6

            total_generation_sum[identifier] = (
                total_generation + total_generation_su
            )

        # Store the results for the current carrier
        results[("total_generation", carrier)] = total_generation_sum

    # Convert results dictionary to a Pandas DataFrame
    # and fill NaN values with 0
    result_filled = pd.DataFrame(results).fillna(0)
    result_filled.index.name = "country_or_node"

    # Total primary renewable energy generation
    total_renewable_energy = result_filled

    # Calculate various other energy imports and primary energy demands
    gas_primary_energy_demand = calculate_primary_energy_demand_gas(
        snapshot_weightings,
        interesting_countries,
        EU27_countries,
        interesting_nodes,
        n_prim_vec,
    )
    oil_primary_energy_demand = calculate_primary_energy_demand_oil(
        snapshot_weightings,
        interesting_countries,
        EU27_countries,
        interesting_nodes,
        n_prim_vec,
    )
    coal_primary_energy_demand = calculate_primary_energy_demand_coal(
        snapshot_weightings,
        interesting_countries,
        EU27_countries,
        interesting_nodes,
        path_industrial_demand,
        mwh_coal_per_mwh_coke,
        n_prim_vec,
    )
    lignite_primary_energy_demand = calculate_primary_energy_demand_lignite(
        snapshot_weightings,
        interesting_countries,
        EU27_countries,
        interesting_nodes,
        n_prim_vec,
    )
    nuclear_primary_energy_demand = calculate_primary_energy_demand_nuclear(
        snapshot_weightings,
        interesting_countries,
        EU27_countries,
        interesting_nodes,
        n_prim_vec,
    )
    biomass_primary_energy_demand = calculate_primary_energy_demand_biomass(
        snapshot_weightings,
        interesting_countries,
        EU27_countries,
        interesting_nodes,
        n_prim_vec,
    )

    # Ensure the indices match by aligning
    # the Series with the DataFrame's index
    gas_primary_energy_demand = gas_primary_energy_demand.reindex(
        result_filled.index, fill_value=0
    )
    oil_primary_energy_demand = oil_primary_energy_demand.reindex(
        result_filled.index, fill_value=0
    )
    coal_primary_energy_demand = coal_primary_energy_demand.reindex(
        result_filled.index, fill_value=0
    )
    lignite_primary_energy_demand = lignite_primary_energy_demand.reindex(
        result_filled.index, fill_value=0
    )
    nuclear_primary_energy_demand = nuclear_primary_energy_demand.reindex(
        result_filled.index, fill_value=0
    )
    biomass_primary_energy_demand = biomass_primary_energy_demand.reindex(
        result_filled.index, fill_value=0
    )

    # Add the values as new columns with specific
    # carrier names to the DataFrame
    result_filled[("gas_primary_energy_demand", "Gas")] = (
        gas_primary_energy_demand
    )
    result_filled[("oil_primary_energy_demand", "Oil")] = (
        oil_primary_energy_demand
    )
    result_filled[("coal_primary_energy_demand", "Coal")] = (
        coal_primary_energy_demand
    )
    result_filled[("lignite_primary_energy_demand", "Lignite")] = (
        lignite_primary_energy_demand
    )
    result_filled[("nuclear_primary_energy_demand", "Nuclear")] = (
        nuclear_primary_energy_demand
    )
    result_filled[("biomass_primary_energy_demand", "Biomass")] = (
        biomass_primary_energy_demand
    )

    # Calculate the total primary energy demand and add it to the DataFrame
    result_filled[("primary_energy_demand", "Total")] = (
        gas_primary_energy_demand
        + oil_primary_energy_demand
        + coal_primary_energy_demand
        + lignite_primary_energy_demand
        + nuclear_primary_energy_demand
        + biomass_primary_energy_demand
        + total_renewable_energy[("total_generation", "solar")]
        + total_renewable_energy[("total_generation", "wind")]
        + total_renewable_energy[("total_generation", "water")]
    )

    return result_filled


def calculate_co2_storage(
    interesting_countries, EU27_countries, interesting_nodes, network
):
    """
    Calculates the total amount of CO2 sequestered for
    specified countries or regions. This function only considers long term storage in sequestration sites.

    Parameters
    ----------
    interesting_countries : list of str
        List of country codes where CO2 storage is to be calculated.
    EU27_countries : list of str
        List of EU27 country codes, used specifically when the EU27
        group is selected.
    interesting_nodes : list of str
        List of node identifiers to include in the CO2 storage
        calculation.
    network : pypsa.Network
        The PyPSA Network object, which contains all data about the
        energy system, including generators, storage units, and other
        infrastructure components.

    Returns
    -------
    pd.Series
        A pandas Series containing the total amounts of CO2 stored and
        sequestered for each country or node, expressed in million
        tonnes (Mt).
    """
    n_co2_store = network.copy()

    # Dictionary to store the Carbon Capture and Storage (CCS) sums for each
    # identifier
    storage_sums = {}

    # Combine countries and nodes in one list for processing
    combined_identifiers = interesting_countries + interesting_nodes

    for identifier in combined_identifiers:
        bus_slice, identifier_list = get_bus_slice(
            EU27_countries,
            identifier,
            interesting_countries,
            interesting_nodes,
        )

        # Filter stores that have 'co2 sequestered' as carrier and belong to
        # the identifier list
        co2_sequestered_index = n_co2_store.stores[
            (n_co2_store.stores.carrier == "co2 sequestered")
            & (n_co2_store.stores.bus.str[bus_slice].isin(identifier_list))
        ].index
        co2_sequestered = n_co2_store.stores.e_nom_opt[
            co2_sequestered_index
        ].sum()

        # Store the result in the dictionary
        storage_sums[identifier] = co2_sequestered.sum() * 1e-6

    # Convert the dictionary into a Pandas Series
    storage_series = pd.Series(storage_sums, name="CCS")
    storage_series.index.name = "country_or_node"

    return storage_series


def calculate_installed_capacity_PHS(
    interesting_countries, EU27_countries, interesting_nodes, network
):
    """
    Calculates the installed capacity of Pumped Hydro Storage (PHS)
    systems for specified countries or regions. This function assesses
    the total capacity of PHS units based on their maximum operational
    hours and nominal power.

    Parameters
    ----------
    interesting_countries : list of str
        List of country codes where the installed capacity is to be
        calculated.
    EU27_countries : list of str
        List of EU27 country codes, used specifically when the EU27
        group is selected.
    interesting_nodes : list of str
        List of node identifiers to include in the capacity
        calculation.
    network : pypsa.Network
        The PyPSA Network object, which contains all data about the
        energy system, including generators, storage units, and other
        infrastructure components.

    Returns
    -------
    pd.Series
        A pandas Series containing the installed capacity of PHS for
        each country or node.
    """
    n_phs = network.copy()
    storage_sums = {}  # Dictionary to store the installed capacity for PHS

    # Combine countries and nodes in one list for processing
    combined_identifiers = interesting_countries + interesting_nodes

    for identifier in combined_identifiers:
        bus_slice, identifier_list = get_bus_slice(
            EU27_countries,
            identifier,
            interesting_countries,
            interesting_nodes,
        )

        # Filter storage units that have 'PHS' as carrier
        # and belong to the identifier list
        phs_index = n_phs.storage_units[
            (n_phs.storage_units.carrier == "PHS")
            & (n_phs.storage_units.bus.str[bus_slice].isin(identifier_list))
        ].index
        PHS_storage = (
            n_phs.storage_units.max_hours[phs_index]
            * n_phs.storage_units.p_nom_opt[phs_index]
        ).sum() * 1e-3

        # Store the result in the dictionary
        storage_sums[identifier] = PHS_storage

    # Convert the dictionary into a Pandas Series
    storage_series = pd.Series(storage_sums, name="phs_storage")
    storage_series.index.name = "country_or_node"

    return storage_series


def calculate_installed_capacity_hydrogen_storage(
    interesting_countries, EU27_countries, interesting_nodes, network
):
    """
    Calculates the installed capacity for hydrogen storage for
    specified countries or regions. This function assesses the total
    storage capacity of facilities specifically designed for hydrogen.

    Parameters
    ----------
    interesting_countries : list of str
        List of country codes where the installed hydrogen storage
        capacity is to be calculated.
    EU27_countries : list of str
        List of EU27 country codes, used specifically when the EU27
        group is selected.
    interesting_nodes : list of str
        List of node identifiers to include in the capacity
        calculation.
    network : pypsa.Network
        The PyPSA Network object, which contains all data about the
        energy system, including storage units and other infrastructure
        components.

    Returns
    -------
    pd.Series
        A pandas Series containing the installed capacity of hydrogen
        storage for each country or node.
    """
    n_h2_store = network.copy()

    # Dictionary to store the installed capacity for hydrogen
    storage_sums = {}

    # Combine countries and nodes in one list for processing
    combined_identifiers = interesting_countries + interesting_nodes

    for identifier in combined_identifiers:
        bus_slice, identifier_list = get_bus_slice(
            EU27_countries,
            identifier,
            interesting_countries,
            interesting_nodes,
        )

        # Filter stores that have 'H2' as carrier and belong
        # to the identifier list
        hydrogen_store_index = n_h2_store.stores[
            (n_h2_store.stores.carrier == "H2")
            & (n_h2_store.stores.bus.str[bus_slice].isin(identifier_list))
        ].index
        hydrogen_store = (
            n_h2_store.stores.e_nom_opt[hydrogen_store_index].sum() * 1e-3
        )

        # Store the result in the dictionary
        storage_sums[identifier] = hydrogen_store

    # Convert the dictionary into a Pandas Series
    storage_series = pd.Series(storage_sums, name="hydrogen_storage")
    storage_series.index.name = "country_or_node"

    return storage_series


def calculate_installed_capacity_electricity_storage(
    interesting_countries, EU27_countries, interesting_nodes, network
):
    """
    Calculates the installed capacity for electricity storage,
    distinguishing between storage for distribution and transmission
    purposes for specified countries or regions. This function
    evaluates the total capacity of different types of electricity
    storage units like lithium-ion batteries, home batteries, and other
    transmission-specific batteries.

    Parameters
    ----------
    interesting_countries : list of str
        List of country codes where the installed electricity storage
        capacity is to be calculated.
    EU27_countries : list of str
        List of EU27 country codes, used specifically when the EU27
        group is selected.
    interesting_nodes : list of str
        List of node identifiers to include in the capacity
        calculation.
    network : pypsa.Network
        The PyPSA Network object, which contains all data about the
        energy system, including storage units and other infrastructure
        components.

    Returns
    -------
    pd.DataFrame
        A pandas DataFrame containing the installed capacities for both
        distribution and transmission electricity storage for each
        country or node.
    """
    n_e_store = network.copy()
    # Dictionary to store the installed capacity for
    # electricity for distribution and transmission separately
    storage_data = {}

    # Combine countries and nodes in one list for processing
    combined_identifiers = interesting_countries + interesting_nodes

    for identifier in combined_identifiers:
        bus_slice, identifier_list = get_bus_slice(
            EU27_countries,
            identifier,
            interesting_countries,
            interesting_nodes,
        )

        distribution_bat = ["Li ion", "home battery"]
        transmission_bat = ["battery"]

        # Filter stores for distribution batteries
        # and calculate their total nominal power
        distribution_storage_index = n_e_store.stores[
            (n_e_store.stores.carrier.isin(distribution_bat))
            & (n_e_store.stores.bus.str[bus_slice].isin(identifier_list))
        ].index
        distribution_storage_capacity = (
            n_e_store.stores.e_nom_opt[distribution_storage_index].sum() * 1e-3
        )

        # Filter stores for transmission batteries
        # and calculate their total nominal power
        transmission_storage_index = n_e_store.stores[
            (n_e_store.stores.carrier.isin(transmission_bat))
            & (n_e_store.stores.bus.str[bus_slice].isin(identifier_list))
        ].index
        transmission_storage_capacity = (
            n_e_store.stores.e_nom_opt[transmission_storage_index].sum() * 1e-3
        )

        # Store the results in the dictionary under each identifier
        storage_data[identifier] = {
            "distribution": distribution_storage_capacity,
            "transmission": transmission_storage_capacity,
        }

    # Convert the dictionary into a Pandas DataFrame
    storage_df = pd.DataFrame.from_dict(storage_data, orient="index")
    storage_df.index.name = "country_or_node"

    return storage_df


def round_dataframe_values(dataframe):
    """
    Rounds the numerical values in a DataFrame to one decimal place and
    replaces all NaN values with 0. This function ensures that the
    DataFrame's numerical data is cleaned and standardized for better
    readability and further analysis.

    Parameters
    ----------
    dataframe : pd.DataFrame
        The pandas DataFrame containing the data to be processed.

    Returns
    -------
    None
    """
    # Replace all NaN values with 0
    dataframe.fillna(0, inplace=True)

    # Iterate through each column in the DataFrame
    for col in dataframe.columns:
        # Round numerical values in each column to one decimal place
        dataframe[col] = dataframe[col].apply(
            lambda x: round(x, 1) if isinstance(x, (int, float)) else x
        )


def import_backup_capacities(
    interesting_countries, interesting_nodes, resultdir, year, scenario
):
    """
    If backup capacities where calculated in the script
    configurable_energy_balances, import them to print them into the
    KPI file.

    Parameters
    ----------
    interesting_countries : list of str
        List of country codes where the installed electricity storage
        capacity is to be calculated.
    interesting_nodes : list of str
        List of node identifiers to include in the analysis.
    resultdir : Path
        Path object with the directory for the KPI table.
    year : int
        List of years to be analyzed.
    scenario : str
        List of scenarios to be analyzed.

    Returns
    -------
    dict
        Dictionary containing the capacities estimated
        post-optimisation for each country/region.
    dict
        Dictionary containing the costs for backup capacities estimated
        post-optimisation for each country/region.
    """
    path_capas = (
        resultdir.parent / "Balances" / "gas_turbine_backup_capacities.xlsx"
    )
    path_costs = (
        resultdir.parent / "Balances" / "gas_turbine_backup_costs.xlsx"
    )
    capas = {}
    costs = {}
    national_scenarios = ["CN", "SN"]
    combined_identifiers = interesting_countries + interesting_nodes
    for identifier in combined_identifiers:
        if identifier in ["All_Countries", "All_Nodes"]:
            sheet_name = "All"
        else:
            sheet_name = identifier
        if (
            os.path.exists(path_capas)
            & (scenario in national_scenarios)
            & (len(identifier) != 5)
        ):
            capas[identifier] = (
                pd.read_excel(
                    path_capas, sheet_name=sheet_name, index_col=0
                ).loc[scenario, year]
                * 1e-3
            )
        else:
            capas[identifier] = 0

        if (
            os.path.exists(path_costs)
            & (scenario in national_scenarios)
            & (len(identifier) != 5)
        ):
            costs[identifier] = (
                pd.read_excel(
                    path_costs, sheet_name=sheet_name, index_col=0
                ).loc[scenario, year]
                * 1e-9
            )
        else:
            costs[identifier] = 0
    return capas, costs


def process_and_compare_networks(
    workbook,
    networks_year,
    framework_path,
    interesting_countries,
    EU27_countries,
    interesting_nodes,
    base_path_industrial_demand,
    base_name_industrial_demand,
    mwh_coal_per_mwh_coke,
    years,
    scenarios,
    resultdir,
):
    """
    Processes and compares network data across multiple scenarios and
    years based on specified parameters related to PyPSA networks. This
    function calculates various energy metrics and populates a provided
    Excel workbook with detailed analytical results for each country or
    node.

    Parameters
    ----------
    workbook : openpyxl.Workbook
        An openpyxl Workbook object to be populated with results.
    networks_year : dict
        The base path to the directory containing the network files.
    framework_path : str
        The path to the framework Excel file that structures the
        analysis.
    interesting_countries : list of str
        List of country codes for which the analysis is performed.
    EU27_countries : list of str
        List of EU27 country codes, used when the EU27 group is
        selected.
    interesting_nodes : list of str
        List of node identifiers to include in the analysis.
    base_path_industrial_demand : str
        The base path to the directory containing industrial demand
        data.
    base_name_industrial_demand : str
        The base name of the industrial demand files.
    mwh_coal_per_mwh_coke : float
        Conversion factor from MWh of coal to MWh of coke.
    years : list of int
        List of years to be analyzed.
    scenarios : list of str
        List of scenarios to be analyzed.
    resultdir : Path
        Path object with the directory for the KPI table.

    Returns
    -------
    None
        This function modifies the provided workbook in place, adding
        sheets with results for each country or node.
    """

    # Load the framework data from the specified path; this will
    # structure the analysis
    framework_df = load_framework(framework_path)

    country_node_data = {}

    # Combine countries and nodes for processing
    all_identifiers = interesting_countries + interesting_nodes

    # Iterate over each year to process network data for each scenario
    for year in years:
        # Construct the path to the industrial demand data for the current year
        path_industrial_demand = (
            f"{base_path_industrial_demand}/"
            f"{base_name_industrial_demand}{year}.csv"
        )

        # Iterate over each scenario, loading the network data
        # from the respective path
        for scenario in scenarios:

            this_network = networks_year[year][scenario]
            # Retrieve the network's time snapshots, used in
            # further calculations
            snapshot_weightings = this_network.snapshot_weightings.generators

            # Calculate various energy metrics for the scenarios
            re_electrictiy = load_and_process_RE_electricity(
                snapshot_weightings,
                interesting_countries,
                EU27_countries,
                interesting_nodes,
                this_network,
            )
            generated_electricity = sum_generated_electricity(
                snapshot_weightings,
                interesting_countries,
                EU27_countries,
                interesting_nodes,
                this_network,
            )
            net_imported_electricity = calculate_net_electricity_import(
                snapshot_weightings,
                interesting_countries,
                EU27_countries,
                interesting_nodes,
                this_network,
            )
            gross_electricity_consumption = (
                calculate_gross_electricity_consumption(
                    snapshot_weightings,
                    interesting_countries,
                    EU27_countries,
                    interesting_nodes,
                    this_network,
                )
            )
            capacity_electricity_interconnectors = (
                calculate_capacity_of_electricity_interconnectors(
                    interesting_countries,
                    EU27_countries,
                    interesting_nodes,
                    this_network,
                )
            )
            total_capacitiy_electricity_links_and_lines = (
                calculate_capacity_of_electricity_links_and_lines(
                    interesting_countries,
                    EU27_countries,
                    interesting_nodes,
                    this_network,
                )
            )

            net_imported_hydrogen = calculate_h2_gas_import(
                snapshot_weightings,
                interesting_countries,
                EU27_countries,
                interesting_nodes,
                this_network,
                "H2",
            )
            generated_hydrogen = calculate_sum_generated_H2(
                snapshot_weightings,
                interesting_countries,
                EU27_countries,
                interesting_nodes,
                this_network,
            )
            gross_hydrogen_consumption = calculate_gross_hydrogen_consumption(
                snapshot_weightings,
                interesting_countries,
                EU27_countries,
                interesting_nodes,
                this_network,
            )
            capacity_H2_interconnectors = calculate_capacity_interconnectors(
                interesting_countries,
                EU27_countries,
                interesting_nodes,
                this_network,
                "H2",
            )
            total_capacitiy_hydrogen_grid = calculate_capacity_grids(
                interesting_countries,
                EU27_countries,
                interesting_nodes,
                this_network,
                "H2",
            )

            capacity_gas_interconnectors = calculate_capacity_interconnectors(
                interesting_countries,
                EU27_countries,
                interesting_nodes,
                this_network,
                "gas",
            )
            total_capacitiy_gas_grid = calculate_capacity_grids(
                interesting_countries,
                EU27_countries,
                interesting_nodes,
                this_network,
                "gas",
            )
            net_imported_methane = calculate_h2_gas_import(
                snapshot_weightings,
                interesting_countries,
                EU27_countries,
                interesting_nodes,
                this_network,
                "gas",
            )

            capacity_co2_interconnectors = calculate_capacity_interconnectors(
                interesting_countries,
                EU27_countries,
                interesting_nodes,
                this_network,
                "CO2",
            )
            total_capacitiy_co2_grid = calculate_capacity_grids(
                interesting_countries,
                EU27_countries,
                interesting_nodes,
                this_network,
                "CO2",
            )

            thermal_capacities, thermal_dispatch = calculate_capacity_thermal(
                interesting_countries,
                EU27_countries,
                interesting_nodes,
                this_network,
            )
            flhs_gas_turbines = calculate_flh_links(
                snapshot_weightings,
                interesting_countries,
                EU27_countries,
                interesting_nodes,
                this_network,
                filter_carrier=["OCGT", "CCGT", "allam"],
            )

            capacity_H2_electrolysis = calculate_capacity_of_H2_electrolysis(
                interesting_countries,
                EU27_countries,
                interesting_nodes,
                this_network,
            )
            capacity_FT_synthesis = calculate_capacity_of_FT_synthesis(
                interesting_countries,
                EU27_countries,
                interesting_nodes,
                this_network,
            )
            generated_energy_H2_electrolysis = (
                calculate_generated_energy_H2_electrolysis(
                    interesting_countries,
                    EU27_countries,
                    interesting_nodes,
                    this_network,
                )
            )
            generated_energy_FT_synthesis = (
                calculate_generated_energy_FT_synthesis(
                    snapshot_weightings,
                    interesting_countries,
                    EU27_countries,
                    interesting_nodes,
                    this_network,
                )
            )
            flhs_electrolysis = calculate_flh_links(
                snapshot_weightings,
                interesting_countries,
                EU27_countries,
                interesting_nodes,
                this_network,
                filter_carrier=["H2 Electrolysis"],
            )

            final_energy_demand_vector = calculate_final_energy_demand_vector(
                snapshot_weightings,
                interesting_countries,
                EU27_countries,
                interesting_nodes,
                path_industrial_demand,
                mwh_coal_per_mwh_coke,
                this_network,
            )
            primary_energy_demand_vector = (
                calculate_primary_energy_demand_vector(
                    snapshot_weightings,
                    interesting_countries,
                    EU27_countries,
                    interesting_nodes,
                    path_industrial_demand,
                    mwh_coal_per_mwh_coke,
                    this_network,
                )
            )

            co2_storage = calculate_co2_storage(
                interesting_countries,
                EU27_countries,
                interesting_nodes,
                this_network,
            )
            PHS_storage_installed_capacity = calculate_installed_capacity_PHS(
                interesting_countries,
                EU27_countries,
                interesting_nodes,
                this_network,
            )
            H2_storage_installed_capacity = (
                calculate_installed_capacity_hydrogen_storage(
                    interesting_countries,
                    EU27_countries,
                    interesting_nodes,
                    this_network,
                )
            )
            Electricity_storage_installed_capacity = (
                calculate_installed_capacity_electricity_storage(
                    interesting_countries,
                    EU27_countries,
                    interesting_nodes,
                    this_network,
                )
            )

            energy_demand_electrification_degree = (
                final_energy_demand_vector["Electrical Energy"]
                / final_energy_demand_vector["Total"]
            )

            backup_capacities, backup_capas_costs = import_backup_capacities(
                interesting_countries,
                interesting_nodes,
                resultdir,
                year,
                scenario,
            )

            for identifier in all_identifiers:
                if identifier not in country_node_data:
                    country_node_data[identifier] = framework_df.copy()

                # Insert RE-ELECTRICITY values into the framework
                for carrier in ["Solar", "Onwind", "Offwind"]:
                    if carrier in re_electrictiy.columns.get_level_values(1):
                        total_generation = re_electrictiy[
                            "total_generation", carrier
                        ].get(identifier, 0)
                        installed_capacity = re_electrictiy[
                            "opt_capacity", carrier
                        ].get(identifier, 0)
                        full_load_hours = re_electrictiy[
                            "full_load_hours", carrier
                        ].get(identifier, 0)
                        curtailment = re_electrictiy[
                            "curtailment", carrier
                        ].get(identifier, 0)
                        # Update the framework DataFrame
                        for i, row in country_node_data[identifier].iterrows():
                            if row["Name"] == carrier:
                                if "Installed capacity" in row["Subcategory"]:
                                    country_node_data[identifier].loc[
                                        i, f"{year} {scenario}"
                                    ] = installed_capacity
                                elif "Generated energy" in row["Subcategory"]:
                                    country_node_data[identifier].loc[
                                        i, f"{year} {scenario}"
                                    ] = total_generation
                                elif "Full-load hours" in row["Subcategory"]:
                                    country_node_data[identifier].loc[
                                        i, f"{year} {scenario}"
                                    ] = full_load_hours
                                elif "Curtailment" in row["Subcategory"]:
                                    country_node_data[identifier].loc[
                                        i, f"{year} {scenario}"
                                    ] = curtailment

                # Insert THERMAL POWER PLANT CAPACITIES into the framework

                for i, row in country_node_data[identifier].iterrows():
                    if (
                        "Methane" in row["Name"]
                        and "Installed capacity" in row["Subcategory"]
                    ):
                        country_node_data[identifier].loc[
                            i, f"{year} {scenario}"
                        ] = thermal_capacities["gas"].get(identifier, 0)

                for i, row in country_node_data[identifier].iterrows():
                    if (
                        "Lignite" in row["Name"]
                        and "Installed capacity" in row["Subcategory"]
                    ):
                        country_node_data[identifier].loc[
                            i, f"{year} {scenario}"
                        ] = thermal_capacities["lignite"].get(identifier, 0)

                for i, row in country_node_data[identifier].iterrows():
                    if (
                        "Hydrogen" in row["Name"]
                        and "Installed capacity" in row["Subcategory"]
                    ):
                        country_node_data[identifier].loc[
                            i, f"{year} {scenario}"
                        ] = thermal_capacities["hydrogen"].get(identifier, 0)

                for i, row in country_node_data[identifier].iterrows():
                    if (
                        "Nuclear" in row["Name"]
                        and "Installed capacity" in row["Subcategory"]
                    ):
                        country_node_data[identifier].loc[
                            i, f"{year} {scenario}"
                        ] = thermal_capacities["nuclear"].get(identifier, 0)

                for i, row in country_node_data[identifier].iterrows():
                    if (
                        "Oil" in row["Name"]
                        and "Installed capacity" in row["Subcategory"]
                    ):
                        country_node_data[identifier].loc[
                            i, f"{year} {scenario}"
                        ] = thermal_capacities["oil"].get(identifier, 0)

                for i, row in country_node_data[identifier].iterrows():
                    if (
                        "Coal" in row["Name"]
                        and "Installed capacity" in row["Subcategory"]
                    ):
                        country_node_data[identifier].loc[
                            i, f"{year} {scenario}"
                        ] = thermal_capacities["coal"].get(identifier, 0)

                # Insert THERMAL POWER PLANT GENERATED ELECTRICITY into
                # the framework

                for i, row in country_node_data[identifier].iterrows():
                    if (
                        "Methane (turbines & CHP)" in row["Name"]
                        and "Generated energy" in row["Subcategory"]
                    ):
                        country_node_data[identifier].loc[
                            i, f"{year} {scenario}"
                        ] = thermal_dispatch["gas"].get(identifier, 0)

                for i, row in country_node_data[identifier].iterrows():
                    if (
                        "Lignite" in row["Name"]
                        and "Generated energy" in row["Subcategory"]
                    ):
                        country_node_data[identifier].loc[
                            i, f"{year} {scenario}"
                        ] = thermal_dispatch["lignite"].get(identifier, 0)

                for i, row in country_node_data[identifier].iterrows():
                    if (
                        "Hydrogen" in row["Name"]
                        and "Generated energy" in row["Subcategory"]
                    ):
                        country_node_data[identifier].loc[
                            i, f"{year} {scenario}"
                        ] = thermal_dispatch["hydrogen"].get(identifier, 0)

                for i, row in country_node_data[identifier].iterrows():
                    if (
                        "Nuclear" in row["Name"]
                        and "Generated energy" in row["Subcategory"]
                    ):
                        country_node_data[identifier].loc[
                            i, f"{year} {scenario}"
                        ] = thermal_dispatch["nuclear"].get(identifier, 0)

                for i, row in country_node_data[identifier].iterrows():
                    if (
                        "Oil" in row["Name"]
                        and "Generated energy" in row["Subcategory"]
                    ):
                        country_node_data[identifier].loc[
                            i, f"{year} {scenario}"
                        ] = thermal_dispatch["oil"].get(identifier, 0)

                for i, row in country_node_data[identifier].iterrows():
                    if (
                        "Coal" in row["Name"]
                        and "Generated energy" in row["Subcategory"]
                    ):
                        country_node_data[identifier].loc[
                            i, f"{year} {scenario}"
                        ] = thermal_dispatch["coal"].get(identifier, 0)

                # Insert FINAL ENERGY DEMAND VECTOR values into the framework

                for i, row in country_node_data[identifier].iterrows():
                    if (
                        "Total" in row["Name"]
                        and "Final energy demand" in row["Subcategory"]
                    ):
                        country_node_data[identifier].loc[
                            i, f"{year} {scenario}"
                        ] = final_energy_demand_vector["Total"].get(
                            identifier, 0
                        )

                for i, row in country_node_data[identifier].iterrows():
                    if (
                        "Hydrogen" in row["Name"]
                        and "Final energy demand" in row["Subcategory"]
                    ):
                        country_node_data[identifier].loc[
                            i, f"{year} {scenario}"
                        ] = final_energy_demand_vector["Hydrogen"].get(
                            identifier, 0
                        )

                for i, row in country_node_data[identifier].iterrows():
                    if (
                        "Heating supply" in row["Name"]
                        and "Final energy demand" in row["Subcategory"]
                    ):
                        country_node_data[identifier].loc[
                            i, f"{year} {scenario}"
                        ] = final_energy_demand_vector["Heat"].get(
                            identifier, 0
                        )

                for i, row in country_node_data[identifier].iterrows():
                    if (
                        "Oil" in row["Name"]
                        and "Final energy demand" in row["Subcategory"]
                    ):
                        country_node_data[identifier].loc[
                            i, f"{year} {scenario}"
                        ] = final_energy_demand_vector["Oil"].get(
                            identifier, 0
                        )

                for i, row in country_node_data[identifier].iterrows():
                    if (
                        "Coal" in row["Name"]
                        and "Final energy demand" in row["Subcategory"]
                    ):
                        country_node_data[identifier].loc[
                            i, f"{year} {scenario}"
                        ] = final_energy_demand_vector["Coal"].get(
                            identifier, 0
                        )

                for i, row in country_node_data[identifier].iterrows():
                    if (
                        "Methane" in row["Name"]
                        and "Final energy demand" in row["Subcategory"]
                    ):
                        country_node_data[identifier].loc[
                            i, f"{year} {scenario}"
                        ] = final_energy_demand_vector["Gas"].get(
                            identifier, 0
                        )

                for i, row in country_node_data[identifier].iterrows():
                    if (
                        "Electrical energy" in row["Name"]
                        and "Final energy demand" in row["Subcategory"]
                    ):
                        country_node_data[identifier].loc[
                            i, f"{year} {scenario}"
                        ] = final_energy_demand_vector[
                            "Electrical Energy"
                        ].get(
                            identifier, 0
                        )

                for i, row in country_node_data[identifier].iterrows():
                    if (
                        "Methanol" in row["Name"]
                        and "Final energy demand" in row["Subcategory"]
                    ):
                        country_node_data[identifier].loc[
                            i, f"{year} {scenario}"
                        ] = final_energy_demand_vector["Methanol"].get(
                            identifier, 0
                        )

                for i, row in country_node_data[identifier].iterrows():
                    if (
                        "Biomass" in row["Name"]
                        and "Final energy demand" in row["Subcategory"]
                    ):
                        country_node_data[identifier].loc[
                            i, f"{year} {scenario}"
                        ] = final_energy_demand_vector["Biomass"].get(
                            identifier, 0
                        )

                for i, row in country_node_data[identifier].iterrows():
                    if (
                        "Households & Services" in row["Name"]
                        and "Final energy demand" in row["Subcategory"]
                    ):
                        country_node_data[identifier].loc[
                            i, f"{year} {scenario}"
                        ] = final_energy_demand_vector[
                            "Households & Services"
                        ].get(
                            identifier, 0
                        )

                for i, row in country_node_data[identifier].iterrows():
                    if (
                        "Agriculture" in row["Name"]
                        and "Final energy demand" in row["Subcategory"]
                    ):
                        country_node_data[identifier].loc[
                            i, f"{year} {scenario}"
                        ] = final_energy_demand_vector["Agriculture"].get(
                            identifier, 0
                        )

                for i, row in country_node_data[identifier].iterrows():
                    if (
                        "Industry" in row["Name"]
                        and "Final energy demand" in row["Subcategory"]
                    ):
                        country_node_data[identifier].loc[
                            i, f"{year} {scenario}"
                        ] = final_energy_demand_vector["Industry"].get(
                            identifier, 0
                        )

                for i, row in country_node_data[identifier].iterrows():
                    if (
                        "Transport" in row["Name"]
                        and "Final energy demand" in row["Subcategory"]
                    ):
                        country_node_data[identifier].loc[
                            i, f"{year} {scenario}"
                        ] = final_energy_demand_vector["Transport"].get(
                            identifier, 0
                        )

                # Insert PRIMARY ENERGY DEMAND VECTOR values into the framework

                for i, row in country_node_data[identifier].iterrows():
                    if (
                        "Wind" in row["Name"]
                        and "Primary energy demand" in row["Subcategory"]
                    ):
                        country_node_data[identifier].loc[
                            i, f"{year} {scenario}"
                        ] = primary_energy_demand_vector[
                            ("total_generation", "wind")
                        ].get(
                            identifier, 0
                        )

                for i, row in country_node_data[identifier].iterrows():
                    if (
                        "Total" in row["Name"]
                        and "Primary energy demand" in row["Subcategory"]
                    ):
                        country_node_data[identifier].loc[
                            i, f"{year} {scenario}"
                        ] = primary_energy_demand_vector[
                            ("primary_energy_demand", "Total")
                        ].get(
                            identifier, 0
                        )

                for i, row in country_node_data[identifier].iterrows():
                    if (
                        "Water" in row["Name"]
                        and "Primary energy demand" in row["Subcategory"]
                    ):
                        country_node_data[identifier].loc[
                            i, f"{year} {scenario}"
                        ] = primary_energy_demand_vector[
                            ("total_generation", "water")
                        ].get(
                            identifier, 0
                        )

                for i, row in country_node_data[identifier].iterrows():
                    if (
                        "Solar" in row["Name"]
                        and "Primary energy demand" in row["Subcategory"]
                    ):
                        country_node_data[identifier].loc[
                            i, f"{year} {scenario}"
                        ] = primary_energy_demand_vector[
                            ("total_generation", "solar")
                        ].get(
                            identifier, 0
                        )

                for i, row in country_node_data[identifier].iterrows():
                    if (
                        "Methane" in row["Name"]
                        and "Primary energy demand" in row["Subcategory"]
                    ):
                        country_node_data[identifier].loc[
                            i, f"{year} {scenario}"
                        ] = primary_energy_demand_vector[
                            ("gas_primary_energy_demand", "Gas")
                        ].get(
                            identifier, 0
                        )

                for i, row in country_node_data[identifier].iterrows():
                    if (
                        "Oil" in row["Name"]
                        and "Primary energy demand" in row["Subcategory"]
                    ):
                        country_node_data[identifier].loc[
                            i, f"{year} {scenario}"
                        ] = primary_energy_demand_vector[
                            ("oil_primary_energy_demand", "Oil")
                        ].get(
                            identifier, 0
                        )

                for i, row in country_node_data[identifier].iterrows():
                    if (
                        "Coal" in row["Name"]
                        and "Primary energy demand" in row["Subcategory"]
                    ):
                        country_node_data[identifier].loc[
                            i, f"{year} {scenario}"
                        ] = primary_energy_demand_vector[
                            ("coal_primary_energy_demand", "Coal")
                        ].get(
                            identifier, 0
                        )

                for i, row in country_node_data[identifier].iterrows():
                    if (
                        "Lignite" in row["Name"]
                        and "Primary energy demand" in row["Subcategory"]
                    ):
                        country_node_data[identifier].loc[
                            i, f"{year} {scenario}"
                        ] = primary_energy_demand_vector[
                            ("lignite_primary_energy_demand", "Lignite")
                        ].get(
                            identifier, 0
                        )

                for i, row in country_node_data[identifier].iterrows():
                    if (
                        "Nuclear" in row["Name"]
                        and "Primary energy demand" in row["Subcategory"]
                    ):
                        country_node_data[identifier].loc[
                            i, f"{year} {scenario}"
                        ] = primary_energy_demand_vector[
                            ("nuclear_primary_energy_demand", "Nuclear")
                        ].get(
                            identifier, 0
                        )

                for i, row in country_node_data[identifier].iterrows():
                    if (
                        "Biomass" in row["Name"]
                        and "Primary energy demand" in row["Subcategory"]
                    ):
                        country_node_data[identifier].loc[
                            i, f"{year} {scenario}"
                        ] = primary_energy_demand_vector[
                            ("biomass_primary_energy_demand", "Biomass")
                        ].get(
                            identifier, 0
                        )

                # Insert KPIs on the transmission grids

                for i, row in country_node_data[identifier].iterrows():
                    if (
                        "Electricity grid" in row["Name"]
                        and "National interconnectors" in row["Subcategory"]
                    ):
                        country_node_data[identifier].loc[
                            i, f"{year} {scenario}"
                        ] = capacity_electricity_interconnectors.get(
                            identifier, 0
                        )

                for i, row in country_node_data[identifier].iterrows():
                    if (
                        "Hydrogen grid" in row["Name"]
                        and "National interconnectors" in row["Subcategory"]
                    ):
                        country_node_data[identifier].loc[
                            i, f"{year} {scenario}"
                        ] = capacity_H2_interconnectors.get(identifier, 0)

                for i, row in country_node_data[identifier].iterrows():
                    if (
                        "Methane grid" in row["Name"]
                        and "National interconnectors" in row["Subcategory"]
                    ):
                        country_node_data[identifier].loc[
                            i, f"{year} {scenario}"
                        ] = capacity_gas_interconnectors.get(identifier, 0)

                for i, row in country_node_data[identifier].iterrows():
                    if (
                        "CO2 grid" in row["Name"]
                        and "National interconnectors" in row["Subcategory"]
                    ):
                        country_node_data[identifier].loc[
                            i, f"{year} {scenario}"
                        ] = capacity_co2_interconnectors.get(identifier, 0)

                for i, row in country_node_data[identifier].iterrows():
                    if (
                        "Electricity grid" in row["Name"]
                        and "Total" in row["Subcategory"]
                    ):
                        country_node_data[identifier].loc[
                            i, f"{year} {scenario}"
                        ] = total_capacitiy_electricity_links_and_lines.get(
                            identifier, 0
                        )

                for i, row in country_node_data[identifier].iterrows():
                    if (
                        "Hydrogen grid" in row["Name"]
                        and "Total" in row["Subcategory"]
                    ):
                        country_node_data[identifier].loc[
                            i, f"{year} {scenario}"
                        ] = total_capacitiy_hydrogen_grid.get(identifier, 0)

                for i, row in country_node_data[identifier].iterrows():
                    if (
                        "Methane grid" in row["Name"]
                        and "Total" in row["Subcategory"]
                    ):
                        country_node_data[identifier].loc[
                            i, f"{year} {scenario}"
                        ] = total_capacitiy_gas_grid.get(identifier, 0)

                for i, row in country_node_data[identifier].iterrows():
                    if (
                        "CO2 grid" in row["Name"]
                        and "Total" in row["Subcategory"]
                    ):
                        country_node_data[identifier].loc[
                            i, f"{year} {scenario}"
                        ] = total_capacitiy_co2_grid.get(identifier, 0)

                # Insert other calculated values into the framework

                # Go through each row of data and update
                # based on the name and/or subcategory

                for i, row in country_node_data[identifier].iterrows():
                    if (
                        "Gas turbines (CCGT, OCGT, Allam cycle)" in row["Name"]
                        and "Full-load hours" in row["Subcategory"]
                    ):
                        country_node_data[identifier].loc[
                            i, f"{year} {scenario}"
                        ] = flhs_gas_turbines.get(identifier, 0)

                for i, row in country_node_data[identifier].iterrows():
                    if (
                        "Degree of electrification of energy demand"
                        in row["Name"]
                    ):
                        country_node_data[identifier].loc[
                            i, f"{year} {scenario}"
                        ] = energy_demand_electrification_degree.get(
                            identifier, 0
                        )

                for i, row in country_node_data[identifier].iterrows():
                    if "Gross electricity generation" in row["Name"]:
                        country_node_data[identifier].loc[
                            i, f"{year} {scenario}"
                        ] = generated_electricity.get(identifier, 0)

                for i, row in country_node_data[identifier].iterrows():
                    if "Net electricity import" in row["Name"]:
                        country_node_data[identifier].loc[
                            i, f"{year} {scenario}"
                        ] = net_imported_electricity.get(identifier, 0)

                for i, row in country_node_data[identifier].iterrows():
                    if "Gross electricity consumption" in row["Name"]:
                        country_node_data[identifier].loc[
                            i, f"{year} {scenario}"
                        ] = gross_electricity_consumption.get(identifier, 0)

                for i, row in country_node_data[identifier].iterrows():
                    if "Gross hydrogen consumption" in row["Name"]:
                        country_node_data[identifier].loc[
                            i, f"{year} {scenario}"
                        ] = gross_hydrogen_consumption.get(identifier, 0)

                for i, row in country_node_data[identifier].iterrows():
                    if "Net hydrogen import" in row["Name"]:
                        country_node_data[identifier].loc[
                            i, f"{year} {scenario}"
                        ] = net_imported_hydrogen.get(identifier, 0)

                for i, row in country_node_data[identifier].iterrows():
                    if "Gross hydrogen generation" in row["Name"]:
                        country_node_data[identifier].loc[
                            i, f"{year} {scenario}"
                        ] = generated_hydrogen.get(identifier, 0)

                for i, row in country_node_data[identifier].iterrows():
                    if "Net methane import" in row["Name"]:
                        country_node_data[identifier].loc[
                            i, f"{year} {scenario}"
                        ] = net_imported_methane.get(identifier, 0)

                for i, row in country_node_data[identifier].iterrows():
                    if "Sequestration" in row["Name"]:
                        country_node_data[identifier].loc[
                            i, f"{year} {scenario}"
                        ] = co2_storage.get(identifier, 0)

                for i, row in country_node_data[identifier].iterrows():
                    if (
                        "Electricity storage (distribution batteries)"
                        in row["Name"]
                    ):
                        country_node_data[identifier].loc[
                            i, f"{year} {scenario}"
                        ] = Electricity_storage_installed_capacity[
                            "distribution"
                        ].get(
                            identifier, 0
                        )

                for i, row in country_node_data[identifier].iterrows():
                    if (
                        "Electricity storage (transmission batteries)"
                        in row["Name"]
                    ):
                        country_node_data[identifier].loc[
                            i, f"{year} {scenario}"
                        ] = Electricity_storage_installed_capacity[
                            "transmission"
                        ].get(
                            identifier, 0
                        )

                for i, row in country_node_data[identifier].iterrows():
                    if "Hydrogen storage" in row["Name"]:
                        country_node_data[identifier].loc[
                            i, f"{year} {scenario}"
                        ] = H2_storage_installed_capacity.get(identifier, 0)

                for i, row in country_node_data[identifier].iterrows():
                    if "Pumped storage power plants" in row["Name"]:
                        country_node_data[identifier].loc[
                            i, f"{year} {scenario}"
                        ] = PHS_storage_installed_capacity.get(identifier, 0)

                for i, row in country_node_data[identifier].iterrows():
                    if (
                        "Electrolysis" in row["Name"]
                        and "Installed capacity" in row["Subcategory"]
                    ):
                        country_node_data[identifier].loc[
                            i, f"{year} {scenario}"
                        ] = capacity_H2_electrolysis.get(identifier, 0)

                for i, row in country_node_data[identifier].iterrows():
                    if (
                        "FT-Synthesis" in row["Name"]
                        and "Installed capacity" in row["Subcategory"]
                    ):
                        country_node_data[identifier].loc[
                            i, f"{year} {scenario}"
                        ] = capacity_FT_synthesis.get(identifier, 0)

                for i, row in country_node_data[identifier].iterrows():
                    if (
                        "Electrolysis" in row["Name"]
                        and "Generated energy" in row["Subcategory"]
                    ):
                        country_node_data[identifier].loc[
                            i, f"{year} {scenario}"
                        ] = generated_energy_H2_electrolysis.get(identifier, 0)

                for i, row in country_node_data[identifier].iterrows():
                    if (
                        "FT-Synthesis" in row["Name"]
                        and "Generated energy" in row["Subcategory"]
                    ):
                        country_node_data[identifier].loc[
                            i, f"{year} {scenario}"
                        ] = generated_energy_FT_synthesis.get(identifier, 0)

                for i, row in country_node_data[identifier].iterrows():
                    if (
                        "Electrolysis" in row["Name"]
                        and "Full-load hours" in row["Subcategory"]
                    ):
                        country_node_data[identifier].loc[
                            i, f"{year} {scenario}"
                        ] = flhs_electrolysis.get(identifier, 0)

                for i, row in country_node_data[identifier].iterrows():
                    if (
                        "Post-optimisation estimate" in row["Name"]
                        and "Annualised investment costs" in row["Subcategory"]
                    ):
                        country_node_data[identifier].loc[
                            i, f"{year} {scenario}"
                        ] = backup_capas_costs[identifier]

                for i, row in country_node_data[identifier].iterrows():
                    if (
                        "Post-optimisation estimate" in row["Name"]
                        and "Installed capacity" in row["Subcategory"]
                    ):
                        country_node_data[identifier].loc[
                            i, f"{year} {scenario}"
                        ] = backup_capacities[identifier]

    # Round values
    for identifier, data in country_node_data.items():
        round_dataframe_values(data)

    # Create sheets for each country and node
    for identifier, data in country_node_data.items():
        sheet_name = identifier
        if sheet_name in workbook.sheetnames:
            del workbook[sheet_name]
        ws = workbook.create_sheet(sheet_name)
        for r in dataframe_to_rows(data, index=False, header=True):
            ws.append(r)


def start_KPI_analysis(networks_year, years, scenarios, main_dir, resultdir):
    """
    Main function to process and compare PyPSA network scenarios
    across multiple years and scenarios.

    The function sets up the paths and parameters for the analysis,
    then calls a series of functions to load, process, and analyze
    network data, ultimately saving the results in an Excel workbook.

    Parameters
    ----------
    networks_year : dict
        Nested dictionary containing PyPSA network objects, indexed by
        year and scenario.
    years : list of int
        List of years to analyze.
    scenarios : list of str
        List of scenarios to analyze.
    main_dir : Path
        Main directory of the project, used to locate input data and
        framework files.
    resultdir : Path
        Path to the directory where the KPI Excel workbook will be
        saved.

    Returns
    -------
    None
    """
    log("Starting: start_KPI_analysis")

    # Directory for the industrial energy demand data
    base_path_industrial_demand = (
        main_dir / "scripts_analysis" / "industrial_energy_demand"
    )
    if not os.path.exists(base_path_industrial_demand):
        base_path_industrial_demand = main_dir / "resources"
    # Filename prefix for the industrial energy demand data
    base_name_industrial_demand = "industrial_energy_demand_elec_s_62_"

    # Directory of the framework file which stores the calculated values
    framework_path = main_dir / "scripts_analysis" / "KPIs-Framework.xlsx"

    # Conversion factor from coal to coke energy content, sourced
    # from `prepare_sector_network.py`
    mwh_coal_per_mwh_coke = 1.366

    # Define lists of countries, regions and nodes
    # 'All_Countries' sums the values of all countries in the list
    # "EU27" sums the values for all EU countries except Cyprus and Malta
    interesting_countries = [
        "All_Countries",
        "EU27",
        "AL",
        "AT",
        "BA",
        "BE",
        "BG",
        "CH",
        "CZ",
        "DE",
        "DK",
        "EE",
        "ES",
        "FI",
        "FR",
        "GB",
        "GR",
        "HR",
        "HU",
        "IE",
        "IT",
        "LT",
        "LU",
        "LV",
        "ME",
        "MK",
        "NL",
        "NO",
        "PL",
        "PT",
        "RO",
        "RS",
        "SE",
        "SI",
        "SK",
    ]
    # EU-countries aside from Cyprus and Malta
    EU27_countries = [
        "DK",
        "LV",
        "NL",
        "IT",
        "PL",
        "DE",
        "FI",
        "BE",
        "AT",
        "CZ",
        "HU",
        "RO",
        "EE",
        "PT",
        "BG",
        "GR",
        "LT",
        "ES",
        "SE",
        "FR",
        "LU",
        "SK",
        "IE",
        "SI",
        "HR",
    ]
    # Interesting nodes for comparison
    # Insert 'All_Nodes' into the list to sum the values of all nodes
    # that in the list
    interesting_nodes = (
        []
    )  # [] # For example: 'All_Nodes', 'IT1 2', 'ES4 0', 'IT1 0', 'FR1 0',
    # 'DE1 2', 'BE1 0', 'DE1 0', 'DE1 1', 'DE1 2', 'DE1 3', 'DE1 4', 'DE1 5',
    # 'DE1 6', 'DE1 7', 'DK1 0', 'DK2 0', 'FR1 0', 'FR1 1', 'FR1 2', 'FR1 3',
    #'FR1 4', 'LU1 0', 'NL1 0', 'NL1 1', 'GB0 0', 'GB0 1', 'GB0 2', 'GB5 0'

    # Track the start time of the script for performance measurement
    start_time = time.time()

    # Initialize a new workbook for the results
    workbook = openpyxl.Workbook()

    # Process data and compare network scenarios across
    # specified countries, regions and nodes
    process_and_compare_networks(
        workbook,
        networks_year,
        framework_path,
        interesting_countries,
        EU27_countries,
        interesting_nodes,
        base_path_industrial_demand,
        base_name_industrial_demand,
        mwh_coal_per_mwh_coke,
        years,
        scenarios,
        resultdir,
    )

    # Remove the default worksheet created with new workbook
    workbook.remove(workbook["Sheet"])

    # Save the compiled workbook
    workbook.save(resultdir / "Summary-Results-KPIs.xlsx")

    # Calculate and display the total runtime of the script
    end_time = time.time()
    print(f"Runtime of the code: {(end_time - start_time) / 60} Minutes")

    log("Done: start_KPI_analysis")


if __name__ == "__main__":
    # User configuration for input data:
    resolution = "2190SEG"  # either 'XXXh' for equal time steps
    # or 'XXXSEG' for time segmentation
    sector_opts = f"elec_s_62_lv99__{resolution}-T-H-B-I-A"
    years = ["2030", "2035", "2040", "2045", "2050"]
    # insert run names for scenarios here:
    runs = {
        "scenario_1": "run_name_1",
        "scenario_2": "run_name_2",
        "scenario_3": "run_name_3",
        "scenario_4": "run_name_4",
    } # analysis on back up capacities are only available for the scenario names of national scenarios "CN" and "SN"

    # User configuration for output data:
    evaluation_name = "analysis_results"
    timestamp = datetime.now().strftime("%Y%m%d")
    full_name = evaluation_name + "_" + timestamp

    # Directories
    main_dir = Path(os.getcwd()).parent
    resultdir = (
        Path(main_dir) / full_name
    )  # folder where Excel files and graphs are stored
    if not os.path.exists(resultdir):
        os.makedirs(resultdir)
    KPI_dir = resultdir / "KPIs"
    if not os.path.exists(KPI_dir):
        os.mkdir(KPI_dir)

    scenarios = list(runs.keys())

    # load all network files from results folder into a nested dictionary
    networks_year = {}
    for year in years:
        curr_networks = {}
        for scen in scenarios:
            run_name = runs[scen]
            path_to_networks = Path(
                f"{main_dir}/results/{run_name}/postnetworks"
            )
            n = pypsa.Network(
                path_to_networks / f"{sector_opts}_{year}.nc"
            ) 
            curr_networks.update({scen: n})

        networks_year.update({year: curr_networks})

    start_KPI_analysis(networks_year, years, scenarios, main_dir, KPI_dir)
